from __future__ import annotations

import csv
import html
import json
import math
import os
import re
import ssl
import threading
import time
import xml.etree.ElementTree as ElementTree
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta
from functools import lru_cache
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pyproj import CRS, Transformer
import shapefile
from shapely.errors import ShapelyError
from shapely.geometry import (
    LineString,
    MultiLineString,
    MultiPoint,
    MultiPolygon,
    Point,
    Polygon,
    box,
    shape as shapely_shape,
)
from shapely.ops import transform as shapely_transform

try:
    import truststore
except ImportError:  # pragma: no cover - dependency is expected in runtime
    truststore = None


class ConnectorConfigurationError(LookupError):
    """Raised when a configured connector cannot be found or parsed."""


class ConnectorExecutionError(RuntimeError):
    """Raised when a connector fails to retrieve payload records."""


@dataclass(slots=True)
class ConnectorCheckpoint:
    checkpoint_ts: datetime | None = None
    checkpoint_cursor: str | None = None
    source_version: str | None = None


@dataclass(slots=True)
class ConnectorFetchPolicy:
    max_attempts: int = 3
    backoff_seconds: float = 0.0
    rate_limit_per_minute: int | None = None
    checkpoint_field: str | None = None
    checkpoint_param: str | None = None


@dataclass(slots=True)
class ConnectorRequestPaginationConfig:
    strategy: str = "none"
    page_size: int = 1000
    max_pages: int | None = None
    parallel_requests: int = 1
    offset_param: str = "resultOffset"
    page_size_param: str = "resultRecordCount"


@dataclass(slots=True)
class ConnectorRequestConfig:
    endpoint_url: str | None = None
    start_urls: list[str] = field(default_factory=list)
    method: str = "GET"
    query_params: dict[str, str] = field(default_factory=dict)
    headers: dict[str, str] = field(default_factory=dict)
    auth_header_name: str | None = None
    auth_env_var: str | None = None
    auth_query_param_name: str | None = None
    record_path: list[str] = field(default_factory=list)
    record_pattern: str | None = None
    body_text: str | None = None
    json_body: dict[str, Any] = field(default_factory=dict)
    timeout_seconds: float = 30.0
    text_encoding: str = "utf-8"
    csv_delimiter: str = ","
    csv_encoding: str = "utf-8"
    zip_member_name: str | None = None
    xlsx_sheet_name: str | None = None
    xlsx_header_row: int = 1
    xlsx_data_start_row: int | None = None
    pagination: ConnectorRequestPaginationConfig = field(
        default_factory=ConnectorRequestPaginationConfig
    )


@dataclass(slots=True)
class SourceConnectorFieldRule:
    target: str
    source: str | list[str] | None = None
    transform: str = "identity"
    template: str | None = None
    default: Any = None
    options: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class SourceConnectorRowFilterRule:
    source: str | list[str] | None = None
    operator: str = "equals"
    value: Any = None
    values: list[Any] = field(default_factory=list)
    options: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ConnectorFetchResult:
    records: list[dict[str, Any]]
    checkpoint_ts: datetime | None
    checkpoint_cursor: str | None
    attempt_count: int


@dataclass(slots=True)
class SourceConnectorDefinition:
    connector_key: str
    source_id: str
    metro_id: str
    interface_name: str
    adapter_type: str
    enabled: bool
    load_strategy: str = "evidence"
    inventory_if_codes: list[str] = field(default_factory=list)
    priority: int = 100
    description: str | None = None
    preprocess_strategy: str | None = None
    preprocess_options: dict[str, Any] = field(default_factory=dict)
    fetch_policy: ConnectorFetchPolicy = field(default_factory=ConnectorFetchPolicy)
    request: ConnectorRequestConfig = field(default_factory=ConnectorRequestConfig)
    field_map: dict[str, str] = field(default_factory=dict)
    field_rules: list[SourceConnectorFieldRule] = field(default_factory=list)
    row_filters: list[SourceConnectorRowFilterRule] = field(default_factory=list)
    static_fields: dict[str, Any] = field(default_factory=dict)
    fixture_records: list[dict[str, Any]] = field(default_factory=list)

    @property
    def normalized_source_id(self) -> str:
        return self.source_id.strip().upper()

    @property
    def normalized_metro_id(self) -> str:
        return self.metro_id.strip().upper()


class _SafeFormatDict(dict[str, Any]):
    def __missing__(self, key: str) -> str:
        return ""


class _RequestRateLimiter:
    def __init__(self) -> None:
        self._last_request_at: dict[str, float] = {}
        self._lock = threading.Lock()

    def wait(self, connector_key: str, rate_limit_per_minute: int | None) -> None:
        if rate_limit_per_minute is None or rate_limit_per_minute <= 0:
            return

        interval_seconds = 60.0 / rate_limit_per_minute
        now = time.monotonic()
        with self._lock:
            last_request_at = self._last_request_at.get(connector_key)
            next_request_at = (
                now
                if last_request_at is None
                else max(now, last_request_at + interval_seconds)
            )
            self._last_request_at[connector_key] = next_request_at

        sleep_seconds = next_request_at - now
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)


_RATE_LIMITER = _RequestRateLimiter()
_TLS_CONTEXT = (
    truststore.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    if truststore is not None
    else None
)


def _open_url(request: Request, timeout: float):
    if _TLS_CONTEXT is None:
        return urlopen(request, timeout=timeout)
    try:
        return urlopen(request, timeout=timeout, context=_TLS_CONTEXT)
    except TypeError as exc:
        if "context" not in str(exc):
            raise
        return urlopen(request, timeout=timeout)


class ConnectorRegistry:
    def __init__(self, definitions: list[SourceConnectorDefinition]) -> None:
        self._definitions_by_connector_key = {
            definition.connector_key: definition for definition in definitions
        }
        self._definitions_by_source_metro: dict[
            tuple[str, str], list[SourceConnectorDefinition]
        ] = {}
        for definition in definitions:
            key = (definition.normalized_source_id, definition.normalized_metro_id)
            self._definitions_by_source_metro.setdefault(key, []).append(definition)

    def get_definition(self, source_id: str, metro_id: str) -> SourceConnectorDefinition:
        key = (source_id.strip().upper(), metro_id.strip().upper())
        candidates = sorted(
            self._definitions_by_source_metro.get(key, []),
            key=lambda definition: (
                not definition.enabled,
                definition.priority,
                definition.connector_key,
            ),
        )
        if not candidates:
            raise ConnectorConfigurationError(
                f"Connector `{key[0]}` for metro `{key[1]}` is not configured."
            )

        enabled_candidates = [definition for definition in candidates if definition.enabled]
        if len(enabled_candidates) == 1:
            return enabled_candidates[0]
        if len(enabled_candidates) > 1:
            top = enabled_candidates[0]
            runner_up = enabled_candidates[1]
            if top.priority != runner_up.priority:
                return top
            raise ConnectorConfigurationError(
                f"Multiple enabled connectors are configured for source `{key[0]}` "
                f"and metro `{key[1]}`."
            )
        return candidates[0]

    def get_definition_by_connector_key(self, connector_key: str) -> SourceConnectorDefinition:
        definition = self._definitions_by_connector_key.get(connector_key.strip())
        if definition is None:
            raise ConnectorConfigurationError(
                f"Connector key `{connector_key.strip()}` is not configured."
            )
        return definition

    def list_definitions(self, *, enabled_only: bool = False) -> list[SourceConnectorDefinition]:
        definitions = list(self._definitions_by_connector_key.values())
        if enabled_only:
            definitions = [definition for definition in definitions if definition.enabled]
        return sorted(
            definitions,
            key=lambda definition: (
                definition.normalized_source_id,
                definition.normalized_metro_id,
                definition.priority,
                definition.connector_key,
            ),
        )


def load_connector_registry(config_path: str) -> ConnectorRegistry:
    path = Path(config_path)
    if not path.exists():
        raise ConnectorConfigurationError(f"Connector config path `{path}` does not exist.")

    payload = json.loads(path.read_text(encoding="utf-8"))
    definitions_payload = payload.get("definitions", [])
    if not isinstance(definitions_payload, list):
        raise ConnectorConfigurationError("Connector config must contain a `definitions` list.")

    definitions = [_parse_definition(item) for item in definitions_payload]
    blueprint_payload = payload.get("inventory_blueprints") or {}
    if bool(blueprint_payload.get("enabled", False)):
        authoritative_inventory_path = (
            _normalize_optional_string(blueprint_payload.get("authoritative_inventory_path"))
            or "configs/authoritative_source_inventory.json"
        )
        inventory_path = Path(authoritative_inventory_path)
        if not inventory_path.is_absolute():
            inventory_path = (path.parent / inventory_path).resolve()
        definitions.extend(
            _build_inventory_blueprint_definitions(
                inventory_path=inventory_path,
                explicit_definitions=definitions,
            )
        )
    return ConnectorRegistry(definitions)


def _build_inventory_blueprint_definitions(
    *,
    inventory_path: Path,
    explicit_definitions: list[SourceConnectorDefinition],
) -> list[SourceConnectorDefinition]:
    if not inventory_path.exists():
        raise ConnectorConfigurationError(
            f"Authoritative inventory path `{inventory_path}` does not exist."
        )

    inventory_payload = json.loads(inventory_path.read_text(encoding="utf-8"))
    sources_payload = inventory_payload.get("sources", [])
    if not isinstance(sources_payload, list):
        raise ConnectorConfigurationError(
            "Authoritative inventory must contain a `sources` list."
        )

    covered_if_codes = {
        if_code
        for definition in explicit_definitions
        for if_code in definition.inventory_if_codes
    }

    blueprints: list[SourceConnectorDefinition] = []
    for source_payload in sources_payload:
        if_code = str(source_payload["if_code"]).strip()
        if if_code in covered_if_codes:
            continue

        adapter_type = _infer_inventory_adapter_type(source_payload)
        metro_id = _infer_inventory_metro_id(source_payload)
        target_table = (
            _normalize_optional_string(source_payload.get("target_table"))
            or "source_evidence"
        )
        load_strategy = _infer_inventory_load_strategy(target_table)
        connector_key = _build_inventory_connector_key(
            if_code=if_code,
            metro_id=metro_id,
            name=str(source_payload["name"]),
        )
        protocol = _normalize_optional_string(source_payload.get("protocol")) or "Unknown"
        request_config = _build_inventory_request_config(
            adapter_type=adapter_type,
            source_payload=source_payload,
        )
        preprocess_strategy = (
            "zoning_overlay_to_parcels"
            if load_strategy == "zoning" and "manual" not in protocol.lower()
            else None
        )

        blueprints.append(
            SourceConnectorDefinition(
                connector_key=connector_key,
                source_id=if_code,
                metro_id=metro_id,
                interface_name=_build_inventory_interface_name(if_code, adapter_type),
                adapter_type=adapter_type,
                enabled=False,
                load_strategy=load_strategy,
                inventory_if_codes=[if_code],
                priority=500 + int(source_payload.get("phase", 1)),
                description=(
                    f"Synthesized blueprint for {if_code} {source_payload['name']} "
                    f"using protocol `{protocol}`."
                ),
                preprocess_strategy=preprocess_strategy,
                fetch_policy=ConnectorFetchPolicy(
                    max_attempts=3,
                    backoff_seconds=1.0,
                    rate_limit_per_minute=30,
                ),
                request=request_config,
            )
        )

    return blueprints


def _infer_inventory_adapter_type(source_payload: dict[str, Any]) -> str:
    protocol = (_normalize_optional_string(source_payload.get("protocol")) or "").lower()
    url = (_normalize_optional_string(source_payload.get("url")) or "").lower()
    if "manual" in protocol or "pdf" in protocol:
        return "manual"
    if "xlsx" in protocol or url.endswith(".xlsx"):
        return "http_xlsx"
    if "shapefile" in protocol:
        return "http_zip_shapefile"
    if "csv" in protocol or "txt" in protocol or "download" in protocol:
        if "rest api" in protocol or "api" in protocol:
            return "http_json"
        return "http_csv"
    if "arcgis" in protocol or "featureserver" in protocol or "mapserver" in protocol:
        return "arcgis_feature_service"
    if "rest api" in protocol or "json" in protocol:
        return "http_json"
    return "manual"


def _infer_inventory_metro_id(source_payload: dict[str, Any]) -> str:
    if_code = str(source_payload["if_code"]).strip().upper()
    city = (_normalize_optional_string(source_payload.get("city")) or "").lower()
    county = (_normalize_optional_string(source_payload.get("county")) or "").lower()
    name = (_normalize_optional_string(source_payload.get("name")) or "").lower()

    if if_code in {"IF-029", "IF-030", "IF-031", "IF-032", "IF-044", "IF-045"}:
        return "DFW"
    if if_code in {"IF-033", "IF-034", "IF-035"} or "houston" in name:
        return "HOU"
    if if_code == "IF-036" or "san antonio" in city or "bexar" in county:
        return "SAT"
    if if_code in {"IF-037", "IF-038", "IF-047"} or "austin" in city or "travis" in county:
        return "AUS"
    if if_code in {"IF-039", "IF-048"} or "el paso" in city or "el paso" in county:
        return "ELP"
    if if_code in {"IF-040", "IF-049"} or "laredo" in city or "webb" in county:
        return "LRD"
    if if_code in {"IF-041", "IF-050"} or "mcallen" in city or "hidalgo" in county:
        return "MFE"
    if if_code in {"IF-042", "IF-051"} or "corpus christi" in city or "nueces" in county:
        return "CRP"
    if if_code == "IF-043" or "midland" in city or "midland" in county:
        return "MAF"
    return "TX"


def _infer_inventory_load_strategy(target_table: str) -> str:
    normalized = target_table.strip().lower()
    if normalized == "raw_parcels":
        return "parcel"
    if normalized == "raw_zoning":
        return "zoning"
    return "evidence"


def _build_inventory_request_config(
    *,
    adapter_type: str,
    source_payload: dict[str, Any],
) -> ConnectorRequestConfig:
    endpoint_url = _normalize_optional_string(source_payload.get("url"))
    if adapter_type == "arcgis_feature_service":
        return ConnectorRequestConfig(
            endpoint_url=endpoint_url,
            query_params={
                "where": "1=1",
                "outFields": "*",
                "returnGeometry": "true",
                "f": "geojson",
                "outSR": "4326",
            },
            timeout_seconds=45.0,
            pagination=ConnectorRequestPaginationConfig(
                strategy="arcgis_offset",
                page_size=1000,
            ),
        )
    if (
        adapter_type == "http_json"
        and str(source_payload.get("if_code")).strip().upper() == "IF-026"
    ):
        return ConnectorRequestConfig(
            endpoint_url=endpoint_url,
            method="POST",
            body_text=(
                "[out:json][timeout:180];"
                "way[\"highway\"](24.2,-106.8,36.8,-93.4);"
                "out tags center;"
            ),
            timeout_seconds=60.0,
        )
    if adapter_type in {"http_csv", "http_zip_csv", "http_zip_shapefile", "http_json"}:
        return ConnectorRequestConfig(
            endpoint_url=endpoint_url,
            timeout_seconds=60.0,
        )
    return ConnectorRequestConfig(endpoint_url=endpoint_url)


def _build_inventory_connector_key(*, if_code: str, metro_id: str, name: str) -> str:
    return f"{_slugify(if_code)}_{_slugify(name)}_{metro_id.lower()}_blueprint"


def _build_inventory_interface_name(if_code: str, adapter_type: str) -> str:
    normalized_adapter = adapter_type.replace("_", "-")
    return f"{if_code.lower()}-{normalized_adapter}-v1"


def fetch_connector_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None = None,
) -> ConnectorFetchResult:
    if not definition.enabled:
        raise ConnectorExecutionError(
            f"Connector `{definition.connector_key}` is disabled and cannot be refreshed."
        )

    if definition.adapter_type == "fixture":
        return _fetch_fixture_records(definition, checkpoint)
    if definition.adapter_type == "http_json":
        return _fetch_http_json_records(definition, checkpoint)
    if definition.adapter_type == "http_html":
        return _fetch_http_html_records(definition, checkpoint)
    if definition.adapter_type == "ercot_dam_spp_html":
        return _fetch_ercot_dam_spp_html_records(definition, checkpoint)
    if definition.adapter_type == "myelisting_html":
        return _fetch_myelisting_html_records(definition, checkpoint)
    if definition.adapter_type == "acrevalue_land_listings":
        return _fetch_acrevalue_land_listing_records(definition, checkpoint)
    if definition.adapter_type == "trueprodigy_public_parcels":
        return _fetch_trueprodigy_public_parcel_records(definition, checkpoint)
    if definition.adapter_type == "http_csv":
        return _fetch_http_csv_records(definition, checkpoint)
    if definition.adapter_type == "http_xlsx":
        return _fetch_http_xlsx_records(definition, checkpoint)
    if definition.adapter_type == "usgs_designmaps_grid":
        return _fetch_usgs_designmaps_grid_records(definition, checkpoint)
    if definition.adapter_type == "http_zip_csv":
        return _fetch_http_zip_csv_records(definition, checkpoint)
    if definition.adapter_type == "http_zip_shapefile":
        return _fetch_http_zip_shapefile_records(definition, checkpoint)
    if definition.adapter_type == "arcgis_feature_service":
        return _fetch_arcgis_feature_service_records(definition, checkpoint)
    if definition.adapter_type == "manual":
        return _fetch_manual_records(definition, checkpoint)

    raise ConnectorConfigurationError(
        f"Connector `{definition.connector_key}` has unsupported adapter type "
        f"`{definition.adapter_type}`."
    )


def _parse_definition(payload: dict[str, Any]) -> SourceConnectorDefinition:
    fetch_policy_payload = payload.get("fetch_policy") or {}
    request_payload = payload.get("request") or {}
    normalized_source_id = str(payload["source_id"]).strip().upper()
    return SourceConnectorDefinition(
        connector_key=str(payload["connector_key"]).strip(),
        source_id=normalized_source_id,
        metro_id=str(payload["metro_id"]).strip().upper(),
        interface_name=str(payload["interface_name"]).strip(),
        adapter_type=str(payload.get("adapter_type", "fixture")).strip().lower(),
        enabled=bool(payload.get("enabled", True)),
        load_strategy=(
            _normalize_optional_string(payload.get("load_strategy"))
            or _default_load_strategy(normalized_source_id)
        ),
        inventory_if_codes=_normalize_string_list(payload.get("inventory_if_codes")),
        priority=int(payload.get("priority", 100)),
        description=_normalize_optional_string(payload.get("description")),
        preprocess_strategy=_normalize_optional_string(payload.get("preprocess_strategy")),
        preprocess_options=dict(payload.get("preprocess_options") or {}),
        fetch_policy=ConnectorFetchPolicy(
            max_attempts=max(int(fetch_policy_payload.get("max_attempts", 3)), 1),
            backoff_seconds=float(fetch_policy_payload.get("backoff_seconds", 0.0)),
            rate_limit_per_minute=(
                int(fetch_policy_payload["rate_limit_per_minute"])
                if fetch_policy_payload.get("rate_limit_per_minute") is not None
                else None
            ),
            checkpoint_field=_normalize_optional_string(
                fetch_policy_payload.get("checkpoint_field")
            ),
            checkpoint_param=_normalize_optional_string(
                fetch_policy_payload.get("checkpoint_param")
            ),
        ),
        request=ConnectorRequestConfig(
            endpoint_url=_normalize_optional_string(request_payload.get("endpoint_url")),
            start_urls=_normalize_string_list(request_payload.get("start_urls")),
            method=_normalize_optional_string(request_payload.get("method")) or "GET",
            query_params={
                str(key): str(value)
                for key, value in (request_payload.get("query_params") or {}).items()
            },
            headers={
                str(key): str(value)
                for key, value in (request_payload.get("headers") or {}).items()
            },
            auth_header_name=_normalize_optional_string(request_payload.get("auth_header_name")),
            auth_env_var=_normalize_optional_string(request_payload.get("auth_env_var")),
            auth_query_param_name=_normalize_optional_string(
                request_payload.get("auth_query_param_name")
            ),
            record_path=_normalize_path(request_payload.get("record_path")),
            record_pattern=_normalize_optional_string(request_payload.get("record_pattern")),
            body_text=_normalize_optional_string(request_payload.get("body_text")),
            json_body=dict(request_payload.get("json_body") or {}),
            timeout_seconds=float(request_payload.get("timeout_seconds", 30.0)),
            text_encoding=(
                _normalize_optional_string(request_payload.get("text_encoding")) or "utf-8"
            ),
            csv_delimiter=(
                _normalize_optional_string(request_payload.get("csv_delimiter")) or ","
            ),
            csv_encoding=(
                _normalize_optional_string(request_payload.get("csv_encoding")) or "utf-8"
            ),
            zip_member_name=_normalize_optional_string(request_payload.get("zip_member_name")),
            xlsx_sheet_name=_normalize_optional_string(
                request_payload.get("xlsx_sheet_name")
            ),
            xlsx_header_row=max(int(request_payload.get("xlsx_header_row", 1)), 1),
            xlsx_data_start_row=(
                max(int(request_payload["xlsx_data_start_row"]), 1)
                if request_payload.get("xlsx_data_start_row") is not None
                else None
            ),
            pagination=ConnectorRequestPaginationConfig(
                strategy=_normalize_optional_string(
                    (request_payload.get("pagination") or {}).get("strategy")
                )
                or "none",
                page_size=max(
                    int((request_payload.get("pagination") or {}).get("page_size", 1000)),
                    1,
                ),
                max_pages=(
                    int((request_payload.get("pagination") or {}).get("max_pages"))
                    if (request_payload.get("pagination") or {}).get("max_pages") is not None
                    else None
                ),
                parallel_requests=max(
                    int(
                        (request_payload.get("pagination") or {}).get(
                            "parallel_requests",
                            1,
                        )
                    ),
                    1,
                ),
                offset_param=_normalize_optional_string(
                    (request_payload.get("pagination") or {}).get("offset_param")
                )
                or "resultOffset",
                page_size_param=_normalize_optional_string(
                    (request_payload.get("pagination") or {}).get("page_size_param")
                )
                or "resultRecordCount",
            ),
        ),
        field_map={
            str(key): str(value)
            for key, value in (payload.get("field_map") or {}).items()
        },
        field_rules=[
            SourceConnectorFieldRule(
                target=str(rule_payload["target"]).strip(),
                source=_normalize_rule_source(rule_payload.get("source")),
                transform=_normalize_optional_string(rule_payload.get("transform")) or "identity",
                template=_normalize_optional_string(rule_payload.get("template")),
                default=rule_payload.get("default"),
                options=dict(rule_payload.get("options") or {}),
            )
            for rule_payload in (payload.get("field_rules") or [])
        ],
        row_filters=[
            SourceConnectorRowFilterRule(
                source=_normalize_rule_source(filter_payload.get("source")),
                operator=_normalize_optional_string(filter_payload.get("operator")) or "equals",
                value=filter_payload.get("value"),
                values=list(filter_payload.get("values") or []),
                options=dict(filter_payload.get("options") or {}),
            )
            for filter_payload in (payload.get("row_filters") or [])
        ],
        static_fields=dict(payload.get("static_fields") or {}),
        fixture_records=[dict(record) for record in payload.get("fixture_records", [])],
    )


def _fetch_fixture_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    filtered_raw_records = _filter_incremental_records(
        definition.fixture_records,
        checkpoint,
        definition.fetch_policy.checkpoint_field,
    )
    filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
    mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
    checkpoint_ts = _max_checkpoint_ts(
        filtered_raw_records,
        definition.fetch_policy.checkpoint_field,
    )
    return ConnectorFetchResult(
        records=mapped_records,
        checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
        checkpoint_cursor=None,
        attempt_count=1,
    )


def _fetch_http_json_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            _RATE_LIMITER.wait(
                definition.connector_key,
                definition.fetch_policy.rate_limit_per_minute,
            )
            request = _build_request(definition, checkpoint)
            with _open_url(request, timeout=definition.request.timeout_seconds) as response:
                payload = json.loads(response.read().decode("utf-8"))

            raw_records = _extract_record_list(payload, definition.request.record_path)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_http_html_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            payload_bytes = _download_connector_bytes(definition, checkpoint)
            raw_records = _parse_html_records(
                payload_bytes,
                encoding=definition.request.text_encoding,
                record_pattern=definition.request.record_pattern,
            )
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError, re.error) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_ercot_dam_spp_html_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            raw_records, checkpoint_ts = _fetch_ercot_dam_spp_rows(definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError, re.error) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_ercot_dam_spp_rows(
    definition: SourceConnectorDefinition,
) -> tuple[list[dict[str, Any]], datetime]:
    if definition.request.endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires `request.endpoint_url`."
        )

    current_url = definition.request.endpoint_url
    current_html = _download_text_from_url(
        current_url,
        definition=definition,
        encoding=definition.request.text_encoding,
    )
    current_rows = _parse_first_html_table_records(current_html)
    if current_rows:
        checkpoint_ts = _ercot_operating_day_to_timestamp(current_rows[0].get("Oper Day"))
        return current_rows, checkpoint_ts

    current_date_value = _extract_html_input_value(current_html, "currentDate")
    if current_date_value is None:
        raise ConnectorExecutionError(
            "ERCOT DAM SPP page did not expose a recoverable `currentDate` value."
        )

    fallback_days = max(int(definition.preprocess_options.get("fallback_days", 1)), 1)
    fallback_operating_day = (
        datetime.strptime(current_date_value, "%m/%d/%Y").date() - timedelta(days=fallback_days)
    )
    dated_url = _build_ercot_dated_report_url(current_url, fallback_operating_day.strftime("%Y%m%d"))
    dated_html = _download_text_from_url(
        dated_url,
        definition=definition,
        encoding=definition.request.text_encoding,
    )
    dated_rows = _parse_first_html_table_records(dated_html)
    if not dated_rows:
        raise ConnectorExecutionError(
            "ERCOT DAM SPP dated report did not contain a settlement-point price table."
        )
    return dated_rows, datetime(
        fallback_operating_day.year,
        fallback_operating_day.month,
        fallback_operating_day.day,
        tzinfo=UTC,
    )


def _ercot_operating_day_to_timestamp(value: Any) -> datetime:
    normalized = _strip_or_none(value)
    if normalized is None:
        return datetime.now(UTC)
    try:
        operating_day = datetime.strptime(normalized, "%m/%d/%Y")
    except ValueError:
        return datetime.now(UTC)
    return operating_day.replace(tzinfo=UTC)


def _build_ercot_dated_report_url(current_url: str, operating_day_yyyymmdd: str) -> str:
    split_url = urlsplit(current_url)
    path_parts = [part for part in split_url.path.split("/") if part]
    if not path_parts:
        raise ConnectorExecutionError("ERCOT DAM SPP URL is missing a path segment.")
    report_name = path_parts[-1]
    dated_path = "/" + "/".join([*path_parts[:-1], f"{operating_day_yyyymmdd}_{report_name}"])
    return urlunsplit(
        (
            split_url.scheme,
            split_url.netloc,
            dated_path,
            split_url.query,
            split_url.fragment,
        )
    )


def _extract_html_input_value(source_text: str, input_id: str) -> str | None:
    match = re.search(
        rf'<input[^>]*id=["\']{re.escape(input_id)}["\'][^>]*value=["\'](?P<value>[^"\']+)["\']',
        source_text,
        re.IGNORECASE,
    )
    if match is not None:
        return _clean_html_text(match.group("value"))
    return None


def _fetch_trueprodigy_public_parcel_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            raw_records = _fetch_trueprodigy_public_parcel_rows(definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or datetime.now(UTC),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (
            ConnectorExecutionError,
            HTTPError,
            URLError,
            TimeoutError,
            ValueError,
            json.JSONDecodeError,
        ) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_trueprodigy_public_parcel_rows(
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    base_url = _require_connector_endpoint(definition)
    auth_context = _build_trueprodigy_auth_context(definition)

    tile_size_degrees = float(definition.preprocess_options.get("tile_size_degrees", 0.05))
    if tile_size_degrees <= 0:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires a positive `tile_size_degrees`."
        )

    max_tiles = definition.preprocess_options.get("max_tiles")
    max_features = definition.preprocess_options.get("max_features")
    pid_batch_size = max(int(definition.preprocess_options.get("pid_batch_size", 500)), 1)

    county_geometry = _fetch_trueprodigy_county_geometry(
        definition,
        base_url=base_url,
        auth_context=auth_context,
    )
    parcel_features = _fetch_trueprodigy_parcel_features(
        definition,
        base_url=base_url,
        auth_context=auth_context,
        county_geometry=county_geometry,
        tile_size_degrees=tile_size_degrees,
        max_tiles=(int(max_tiles) if max_tiles is not None else None),
        max_features=(int(max_features) if max_features is not None else None),
    )
    if not parcel_features:
        return []

    appraisal_fields_by_pid = _fetch_trueprodigy_appraisal_fields(
        definition,
        base_url=base_url,
        auth_context=auth_context,
        parcel_ids=list(parcel_features.keys()),
        pid_batch_size=pid_batch_size,
    )

    raw_records: list[dict[str, Any]] = []
    for parcel_id in sorted(parcel_features, key=lambda value: int(str(value)) if str(value).isdigit() else str(value)):
        feature = parcel_features[parcel_id]
        properties = dict(feature.get("properties") or {})
        record = dict(properties)
        record["pID"] = parcel_id
        record["pid"] = parcel_id
        record["county"] = _strip_or_none(properties.get("county")) or auth_context.office_name
        appraisal_fields = appraisal_fields_by_pid.get(parcel_id)
        if appraisal_fields is not None:
            record.update(appraisal_fields)
        geometry_payload = feature.get("geometry")
        if isinstance(geometry_payload, dict):
            record["__geometry__"] = geometry_payload
        raw_records.append(record)

    return raw_records


def _require_connector_endpoint(definition: SourceConnectorDefinition) -> str:
    endpoint_url = _normalize_optional_string(definition.request.endpoint_url)
    if endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires `request.endpoint_url`."
        )
    return endpoint_url.rstrip("/")


@dataclass(slots=True)
class _TrueProdigyAuthContext:
    token_endpoint_url: str
    office_name: str
    token: str | None = None
    token_lock: Any = field(default_factory=threading.Lock, repr=False)

    def get_token(self, definition: SourceConnectorDefinition) -> str:
        with self.token_lock:
            if self.token is None:
                self.token = _fetch_trueprodigy_public_token(
                    definition,
                    token_endpoint_url=self.token_endpoint_url,
                    office_name=self.office_name,
                )
            return self.token

    def refresh_token(self, definition: SourceConnectorDefinition) -> str:
        with self.token_lock:
            self.token = _fetch_trueprodigy_public_token(
                definition,
                token_endpoint_url=self.token_endpoint_url,
                office_name=self.office_name,
            )
            return self.token


def _build_trueprodigy_auth_context(
    definition: SourceConnectorDefinition,
) -> _TrueProdigyAuthContext:
    office_name = (
        _normalize_optional_string(definition.preprocess_options.get("office"))
        or _normalize_optional_string(definition.request.query_params.get("office"))
    )
    if office_name is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires preprocess option `office`."
        )

    token_endpoint_url = _normalize_optional_string(
        definition.preprocess_options.get("token_endpoint_url")
    )
    if token_endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires preprocess option "
            "`token_endpoint_url`."
        )

    return _TrueProdigyAuthContext(
        token_endpoint_url=token_endpoint_url,
        office_name=office_name,
    )


def _fetch_trueprodigy_public_token(
    definition: SourceConnectorDefinition,
    *,
    token_endpoint_url: str,
    office_name: str,
) -> str:
    _RATE_LIMITER.wait(
        definition.connector_key,
        definition.fetch_policy.rate_limit_per_minute,
    )
    request = Request(
        url=token_endpoint_url,
        headers={
            "Content-Type": "application/json",
            "Cache-Control": "no-cache",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "DenseDataCenterLocator/1.0 (+https://github.com/v1kz-ui/Data-Center-Services)",
        },
        data=json.dumps({"office": office_name}).encode("utf-8"),
        method="POST",
    )
    with _open_url(request, timeout=definition.request.timeout_seconds) as response:
        payload = json.loads(response.read().decode("utf-8"))

    token = _resolve_field(payload, "user.token")
    normalized_token = _strip_or_none(token)
    if normalized_token is None:
        raise ConnectorExecutionError(
            f"True Prodigy auth response for `{definition.connector_key}` did not include "
            "`user.token`."
        )
    return normalized_token


def _fetch_trueprodigy_county_geometry(
    definition: SourceConnectorDefinition,
    *,
    base_url: str,
    auth_context: _TrueProdigyAuthContext,
):
    payload = _fetch_trueprodigy_json(
        definition,
        url=f"{base_url}/gama/countylines/geojson",
        auth_context=auth_context,
    )
    feature_collection = _find_trueprodigy_feature_collection(payload)
    if feature_collection is None:
        raise ConnectorExecutionError(
            f"Connector `{definition.connector_key}` did not return county boundary geometry."
        )

    features = feature_collection.get("features")
    if not isinstance(features, list) or not features:
        raise ConnectorExecutionError(
            f"Connector `{definition.connector_key}` returned an empty county boundary payload."
        )

    polygons = [
        shapely_shape(feature.get("geometry"))
        for feature in features
        if isinstance(feature, dict) and isinstance(feature.get("geometry"), dict)
    ]
    if not polygons:
        raise ConnectorExecutionError(
            f"Connector `{definition.connector_key}` did not return usable boundary polygons."
        )

    county_geometry = polygons[0]
    for polygon in polygons[1:]:
        county_geometry = county_geometry.union(polygon)
    return county_geometry


def _fetch_trueprodigy_parcel_features(
    definition: SourceConnectorDefinition,
    *,
    base_url: str,
    auth_context: _TrueProdigyAuthContext,
    county_geometry,
    tile_size_degrees: float,
    max_tiles: int | None,
    max_features: int | None,
) -> dict[str, dict[str, Any]]:
    parcel_features: dict[str, dict[str, Any]] = {}
    tile_geometries = _build_trueprodigy_tile_geometries(
        county_geometry=county_geometry,
        tile_size_degrees=tile_size_degrees,
        max_tiles=max_tiles,
    )
    parallel_requests = _trueprodigy_parallel_requests(definition)

    def ingest_feature_collection(feature_collection: dict[str, Any]) -> bool:
        for feature in feature_collection.get("features", []):
            if not isinstance(feature, dict):
                continue
            properties = feature.get("properties")
            if not isinstance(properties, dict):
                continue
            parcel_id = _strip_or_none(properties.get("pid")) or _strip_or_none(feature.get("id"))
            if parcel_id is None or parcel_id in parcel_features:
                continue
            geometry_payload = feature.get("geometry")
            if not isinstance(geometry_payload, dict):
                continue
            parcel_features[parcel_id] = {
                "id": parcel_id,
                "type": feature.get("type", "Feature"),
                "geometry": geometry_payload,
                "properties": properties,
            }
            if max_features is not None and len(parcel_features) >= max_features:
                return True
        return False

    if parallel_requests <= 1:
        for tile_geometry in tile_geometries:
            feature_collection = _fetch_trueprodigy_parcel_tile(
                definition,
                base_url=base_url,
                auth_context=auth_context,
                tile_geometry=tile_geometry,
            )
            if ingest_feature_collection(feature_collection):
                return parcel_features
        return parcel_features

    with ThreadPoolExecutor(max_workers=parallel_requests) as executor:
        futures = [
            executor.submit(
                _fetch_trueprodigy_parcel_tile,
                definition,
                base_url=base_url,
                auth_context=auth_context,
                tile_geometry=tile_geometry,
            )
            for tile_geometry in tile_geometries
        ]
        for future in as_completed(futures):
            if ingest_feature_collection(future.result()):
                return parcel_features

    return parcel_features


def _build_trueprodigy_tile_geometries(
    *,
    county_geometry,
    tile_size_degrees: float,
    max_tiles: int | None,
) -> list[Any]:
    min_x, min_y, max_x, max_y = county_geometry.bounds
    tile_geometries: list[Any] = []

    x_cursor = min_x
    while x_cursor < max_x:
        next_x = min(x_cursor + tile_size_degrees, max_x)
        y_cursor = min_y
        while y_cursor < max_y:
            next_y = min(y_cursor + tile_size_degrees, max_y)
            tile_geometry = box(x_cursor, y_cursor, next_x, next_y)
            y_cursor = next_y
            if not tile_geometry.intersects(county_geometry):
                continue

            tile_geometries.append(tile_geometry)
            if max_tiles is not None and len(tile_geometries) >= max_tiles:
                return tile_geometries
        x_cursor = next_x

    return tile_geometries


def _trueprodigy_parallel_requests(definition: SourceConnectorDefinition) -> int:
    raw_value = definition.preprocess_options.get("parallel_requests", 1)
    try:
        requested = int(raw_value)
    except (TypeError, ValueError):
        requested = 1
    return max(1, min(requested, 16))


def _fetch_trueprodigy_parcel_tile(
    definition: SourceConnectorDefinition,
    *,
    base_url: str,
    auth_context: _TrueProdigyAuthContext,
    tile_geometry,
) -> dict[str, Any]:
    ring = [
        [tile_geometry.bounds[0], tile_geometry.bounds[1]],
        [tile_geometry.bounds[2], tile_geometry.bounds[1]],
        [tile_geometry.bounds[2], tile_geometry.bounds[3]],
        [tile_geometry.bounds[0], tile_geometry.bounds[3]],
        [tile_geometry.bounds[0], tile_geometry.bounds[1]],
    ]
    points = quote(json.dumps(ring, separators=(",", ":")), safe="")
    payload = _fetch_trueprodigy_json(
        definition,
        url=f"{base_url}/gama/parcelswithinbounds?points={points}",
        auth_context=auth_context,
    )
    feature_collection = _find_trueprodigy_feature_collection(payload)
    if feature_collection is None:
        return {"type": "FeatureCollection", "features": []}
    return feature_collection


def _fetch_trueprodigy_appraisal_fields(
    definition: SourceConnectorDefinition,
    *,
    base_url: str,
    auth_context: _TrueProdigyAuthContext,
    parcel_ids: list[str],
    pid_batch_size: int,
) -> dict[str, dict[str, Any]]:
    fields_by_pid: dict[str, dict[str, Any]] = {}
    batches = [
        parcel_ids[start_index : start_index + pid_batch_size]
        for start_index in range(0, len(parcel_ids), pid_batch_size)
    ]
    parallel_requests = _trueprodigy_parallel_requests(definition)

    if parallel_requests <= 1:
        payloads = [
            _fetch_trueprodigy_appraisal_batch(
                definition,
                base_url=base_url,
                auth_context=auth_context,
                batch=batch,
            )
            for batch in batches
        ]
    else:
        with ThreadPoolExecutor(max_workers=parallel_requests) as executor:
            futures = [
                executor.submit(
                    _fetch_trueprodigy_appraisal_batch,
                    definition,
                    base_url=base_url,
                    auth_context=auth_context,
                    batch=batch,
                )
                for batch in batches
            ]
            payloads = [future.result() for future in as_completed(futures)]

    for payload in payloads:
        results = payload.get("results")
        if not isinstance(results, list):
            continue
        for result in results:
            if not isinstance(result, dict):
                continue
            parcel_id = _strip_or_none(result.get("pID")) or _strip_or_none(result.get("pid"))
            if parcel_id is None:
                continue
            fields_by_pid[parcel_id] = dict(result)

    return fields_by_pid


def _fetch_trueprodigy_appraisal_batch(
    definition: SourceConnectorDefinition,
    *,
    base_url: str,
    auth_context: _TrueProdigyAuthContext,
    batch: list[str],
) -> dict[str, Any]:
    return _fetch_trueprodigy_json(
        definition,
        url=f"{base_url}/gama/appraisalfields/public",
        auth_context=auth_context,
        method="POST",
        json_body={
            "pIDList": [
                int(parcel_id) if parcel_id.isdigit() else parcel_id for parcel_id in batch
            ]
        },
    )


def _fetch_trueprodigy_json(
    definition: SourceConnectorDefinition,
    *,
    url: str,
    auth_context: _TrueProdigyAuthContext,
    method: str = "GET",
    json_body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    _RATE_LIMITER.wait(
        definition.connector_key,
        definition.fetch_policy.rate_limit_per_minute,
    )
    data: bytes | None = None
    if json_body is not None:
        data = json.dumps(json_body).encode("utf-8")

    for auth_attempt in range(2):
        headers = {
            "Authorization": auth_context.get_token(definition),
            "Cache-Control": "no-cache",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "DenseDataCenterLocator/1.0 (+https://github.com/v1kz-ui/Data-Center-Services)",
        }
        if json_body is not None:
            headers["Content-Type"] = "application/json"

        request = Request(url=url, headers=headers, data=data, method=method)
        try:
            with _open_url(request, timeout=definition.request.timeout_seconds) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            if exc.code == 401 and auth_attempt == 0:
                auth_context.refresh_token(definition)
                continue
            raise

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed to retrieve True Prodigy payload "
        f"from `{url}` after refreshing the public token."
    )


def _find_trueprodigy_feature_collection(payload: Any) -> dict[str, Any] | None:
    if isinstance(payload, dict):
        if payload.get("type") == "FeatureCollection" and isinstance(payload.get("features"), list):
            return payload
        for value in payload.values():
            feature_collection = _find_trueprodigy_feature_collection(value)
            if feature_collection is not None:
                return feature_collection
        return None

    if isinstance(payload, list):
        for item in payload:
            feature_collection = _find_trueprodigy_feature_collection(item)
            if feature_collection is not None:
                return feature_collection
    return None


def _fetch_usgs_designmaps_grid_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            raw_records = _fetch_usgs_designmaps_grid_rows(definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=datetime.now(UTC),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (
            HTTPError,
            URLError,
            TimeoutError,
            ValueError,
            FileNotFoundError,
            json.JSONDecodeError,
        ) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_usgs_designmaps_grid_rows(definition: SourceConnectorDefinition) -> list[dict[str, Any]]:
    endpoint_url = _normalize_optional_string(definition.request.endpoint_url)
    if endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires `request.endpoint_url`."
        )

    boundary_geojson_path = _normalize_optional_string(
        definition.preprocess_options.get("boundary_geojson_path")
    )
    if boundary_geojson_path is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires `boundary_geojson_path`."
        )

    boundary_geometry = _load_local_geojson_geometry(boundary_geojson_path)
    grid_step_degrees = float(definition.preprocess_options.get("grid_step_degrees", 0.5))
    if grid_step_degrees <= 0:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires a positive `grid_step_degrees`."
        )

    query_params = {
        key: value
        for key, value in definition.request.query_params.items()
        if _normalize_optional_string(value) is not None
    }
    candidate_points = list(_iter_grid_points_within_geometry(boundary_geometry, grid_step_degrees))
    parallel_requests = max(int(definition.preprocess_options.get("parallel_requests", 8)), 1)

    raw_records: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=parallel_requests) as executor:
        futures = [
            executor.submit(
                _fetch_usgs_designmaps_grid_point,
                definition=definition,
                endpoint_url=endpoint_url,
                query_params=query_params,
                latitude=latitude,
                longitude=longitude,
            )
            for latitude, longitude in candidate_points
        ]
        for future in as_completed(futures):
            raw_records.append(future.result())

    return sorted(raw_records, key=lambda record: record["point_id"])


def _fetch_usgs_designmaps_grid_point(
    *,
    definition: SourceConnectorDefinition,
    endpoint_url: str,
    query_params: dict[str, str],
    latitude: float,
    longitude: float,
) -> dict[str, Any]:
    latitude_text = f"{latitude:.6f}"
    longitude_text = f"{longitude:.6f}"
    title = f"TX Grid {latitude_text},{longitude_text}"
    url = _append_query_params(
        endpoint_url,
        {
            **query_params,
            "latitude": latitude_text,
            "longitude": longitude_text,
            "title": title,
        },
    )
    payload = json.loads(_download_url_bytes(url, definition=definition).decode("utf-8"))
    data = payload.get("response", {}).get("data")
    if not isinstance(data, dict):
        raise ConnectorExecutionError(
            "USGS design maps response did not contain `response.data`."
        )

    point_id = f"{latitude_text},{longitude_text}"
    return {
        "point_id": point_id,
        "latitude": latitude_text,
        "longitude": longitude_text,
        "reference_document": payload.get("request", {}).get("referenceDocument"),
        "risk_category": query_params.get("riskCategory"),
        "site_class": query_params.get("siteClass"),
        "pgam": data.get("pgam"),
        "ss": data.get("ss"),
        "s1": data.get("s1"),
        "sms": data.get("sms"),
        "sm1": data.get("sm1"),
        "sds": data.get("sds"),
        "sd1": data.get("sd1"),
        "sdc": data.get("sdc"),
        "ts": data.get("ts"),
        "t0": data.get("t0"),
        "tl": data.get("tl"),
    }


def _fetch_myelisting_html_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            raw_records = _fetch_myelisting_search_pages(definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or datetime.now(UTC),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (
            ConnectorExecutionError,
            HTTPError,
            URLError,
            TimeoutError,
            ValueError,
            json.JSONDecodeError,
            re.error,
        ) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_acrevalue_land_listing_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            raw_records = _fetch_acrevalue_land_listing_pages(definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or datetime.now(UTC),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (
            ConnectorExecutionError,
            HTTPError,
            URLError,
            TimeoutError,
            ValueError,
            json.JSONDecodeError,
        ) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_acrevalue_land_listing_pages(
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    page_size = max(1, definition.request.pagination.page_size)
    first_page = _fetch_acrevalue_land_listing_page(
        definition,
        page_number=1,
        page_size=page_size,
    )
    first_page_records, listing_count = first_page
    if not first_page_records:
        return []

    total_pages = max(math.ceil(listing_count / page_size), 1)
    if definition.request.pagination.max_pages is not None:
        total_pages = min(total_pages, definition.request.pagination.max_pages)
    if total_pages <= 1:
        return _dedupe_acrevalue_land_listing_records(first_page_records)

    remaining_page_numbers = list(range(2, total_pages + 1))
    collected_records = list(first_page_records)
    max_workers = max(1, min(16, definition.request.pagination.parallel_requests))
    if max_workers <= 1:
        for page_number in remaining_page_numbers:
            page_records, _listing_count = _fetch_acrevalue_land_listing_page(
                definition,
                page_number=page_number,
                page_size=page_size,
            )
            collected_records.extend(page_records)
        return _dedupe_acrevalue_land_listing_records(collected_records)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(
                _fetch_acrevalue_land_listing_page,
                definition,
                page_number=page_number,
                page_size=page_size,
            ): page_number
            for page_number in remaining_page_numbers
        }
        for future in as_completed(futures):
            page_records, _listing_count = future.result()
            collected_records.extend(page_records)

    return _dedupe_acrevalue_land_listing_records(collected_records)


def _fetch_acrevalue_land_listing_page(
    definition: SourceConnectorDefinition,
    *,
    page_number: int,
    page_size: int,
    allow_fallback: bool = True,
) -> tuple[list[dict[str, Any]], int]:
    page_url = _build_acrevalue_land_listing_page_url(
        definition,
        page_number=page_number,
        page_size=page_size,
    )
    last_exception: Exception | None = None
    for attempt_index in range(1, 4):
        try:
            payload = json.loads(
                _download_text_from_url(
                    page_url,
                    definition=definition,
                    encoding=definition.request.text_encoding,
                )
            )
            return _extract_acrevalue_land_listing_page_payload(payload)
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
            last_exception = exc
            if attempt_index < 3:
                time.sleep(min(definition.fetch_policy.backoff_seconds, 2.0))

    if allow_fallback and page_size > 100 and _is_retryable_http_exception(last_exception):
        return _fetch_acrevalue_land_listing_page_fallback(
            definition,
            page_number=page_number,
            page_size=page_size,
        )

    if isinstance(last_exception, Exception):
        raise last_exception
    raise ConnectorExecutionError(f"AcreValue page {page_number} did not return a payload.")


def _fetch_acrevalue_land_listing_page_fallback(
    definition: SourceConnectorDefinition,
    *,
    page_number: int,
    page_size: int,
) -> tuple[list[dict[str, Any]], int]:
    fallback_page_size = 100
    start_offset = (page_number - 1) * page_size
    end_offset = start_offset + page_size
    start_page = (start_offset // fallback_page_size) + 1
    end_page = math.ceil(end_offset / fallback_page_size)
    collected_records: list[dict[str, Any]] = []
    listing_count = 0
    for fallback_page_number in range(start_page, end_page + 1):
        page_records, page_listing_count = _fetch_acrevalue_land_listing_page(
            definition,
            page_number=fallback_page_number,
            page_size=fallback_page_size,
            allow_fallback=False,
        )
        listing_count = max(listing_count, page_listing_count)
        collected_records.extend(page_records)
    return collected_records, listing_count or len(collected_records)


def _extract_acrevalue_land_listing_page_payload(
    payload: Any,
) -> tuple[list[dict[str, Any]], int]:
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        raise ConnectorExecutionError("AcreValue listing response did not contain `data`.")

    listings = data.get("listings")
    if not isinstance(listings, list):
        raise ConnectorExecutionError(
            "AcreValue listing response did not contain `data.listings`."
        )

    listing_count = _normalize_acrevalue_listing_count(
        data.get("listings_count"),
        fallback=len(listings),
    )
    normalized_records = [
        record
        for listing in listings
        if isinstance(listing, dict)
        if (record := _normalize_acrevalue_land_listing(listing)) is not None
    ]
    return normalized_records, listing_count


def _is_retryable_http_exception(exc: Exception | None) -> bool:
    if isinstance(exc, HTTPError):
        return exc.code in {500, 502, 503, 504}
    return isinstance(exc, (URLError, TimeoutError, json.JSONDecodeError))


def _build_acrevalue_land_listing_page_url(
    definition: SourceConnectorDefinition,
    *,
    page_number: int,
    page_size: int,
) -> str:
    endpoint_url = _normalize_optional_string(definition.request.endpoint_url)
    if endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` requires `request.endpoint_url`."
        )
    query_params = {
        **definition.request.query_params,
        "data_limit": str(page_size),
        "page": str(page_number),
    }
    return _append_query_params(endpoint_url, query_params)


def _normalize_acrevalue_listing_count(value: Any, *, fallback: int) -> int:
    try:
        listing_count = int(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return fallback
    return max(listing_count, fallback)


def _normalize_acrevalue_land_listing(listing: dict[str, Any]) -> dict[str, Any] | None:
    source_listing_key = _strip_or_none(listing.get("listing_id")) or _strip_or_none(
        listing.get("guid")
    )
    source_url = _normalize_public_listing_url(
        _strip_or_none(listing.get("standalone_link"))
        or _strip_or_none(listing.get("link")),
        base_url="https://www.acrevalue.com",
    )
    if source_listing_key is None or source_url is None:
        return None

    listing_title = (
        _clean_html_text(listing.get("listing_title"))
        or _derive_acrevalue_listing_title(listing)
        or source_listing_key
    )
    broker_name = _first_clean_acrevalue_listing_value(
        listing,
        "broker",
        "found_broker_name",
        "broker_name",
        "agent",
    )
    listing_refreshed = _normalize_acrevalue_date(
        listing.get("date_updated")
        or listing.get("date_listed")
        or listing.get("date_created")
    )
    return {
        "source_listing_key": source_listing_key,
        "listing_title": listing_title,
        "listing_status": _normalize_acrevalue_listing_status(
            listing.get("status_classification") or listing.get("status")
        ),
        "asset_type": _first_clean_acrevalue_listing_value(
            listing,
            "property_types",
            "property_classification",
            "property_type_one",
            "property_type_two",
            "property_type_three",
        ),
        "asking_price": _normalize_positive_numeric_string(listing.get("price")),
        "acreage": _normalize_positive_numeric_string(listing.get("acres")),
        "building_sqft": None,
        "address_line1": _clean_html_text(listing.get("address")),
        "city": _clean_html_text(listing.get("city")),
        "state_code": _strip_or_none(listing.get("state")),
        "postal_code": _strip_or_none(listing.get("zipcode")),
        "latitude": _normalize_decimal_string(listing.get("latitude")),
        "longitude": _normalize_decimal_string(listing.get("longitude")),
        "broker_name": broker_name,
        "source_url": source_url,
        "listing_refreshed": listing_refreshed,
        "last_verified_at": listing_refreshed,
        "lineage_key": f"listing:acrevalue:{source_listing_key}",
    }


def _dedupe_acrevalue_land_listing_records(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    deduped_records: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    for record in records:
        source_listing_key = _strip_or_none(record.get("source_listing_key"))
        if source_listing_key is None or source_listing_key in seen_keys:
            continue
        seen_keys.add(source_listing_key)
        deduped_records.append(record)
    return deduped_records


def _fetch_myelisting_search_pages(
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    requested_status = _infer_myelisting_requested_status(definition)
    discovered_offers = _discover_myelisting_offers(
        definition,
        requested_status=requested_status,
    )
    if not discovered_offers:
        return []

    max_workers = _determine_myelisting_worker_count(definition)
    if max_workers <= 1:
        return [
            record
            for offer in discovered_offers
            if (
                record := _fetch_myelisting_record_from_offer(
                    offer,
                    definition=definition,
                    requested_status=requested_status,
                )
            )
            is not None
        ]

    collected_records: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(
                _fetch_myelisting_record_from_offer,
                offer,
                definition=definition,
                requested_status=requested_status,
            )
            for offer in discovered_offers
        ]
        for future in as_completed(futures):
            record = future.result()
            if record is not None:
                collected_records.append(record)

    return collected_records


def _fetch_myelisting_record_from_offer(
    offer: dict[str, Any],
    *,
    definition: SourceConnectorDefinition,
    requested_status: str | None,
) -> dict[str, Any] | None:
    detail_url = _strip_or_none(offer.get("url"))
    if detail_url is None:
        return None

    detail_html = _download_text_from_url(
        detail_url,
        definition=definition,
        encoding=definition.request.text_encoding,
    )
    record = _extract_myelisting_detail_record(
        detail_html,
        detail_url=detail_url,
        offer=offer,
        page_url=_strip_or_none(offer.get("page_url")) or detail_url,
    )
    if requested_status is not None and record.get("listing_status") != requested_status:
        return None
    return record


def _determine_myelisting_worker_count(definition: SourceConnectorDefinition) -> int:
    configured_parallelism = definition.request.pagination.parallel_requests
    if configured_parallelism > 1:
        return max(2, min(24, configured_parallelism))

    rate_limit = definition.fetch_policy.rate_limit_per_minute
    if rate_limit is None or rate_limit <= 0:
        return 4
    return max(2, min(12, rate_limit // 30 or 1))


def _discover_myelisting_offers(
    definition: SourceConnectorDefinition,
    *,
    requested_status: str | None,
) -> list[dict[str, Any]]:
    start_urls = definition.request.start_urls or []
    if not start_urls:
        endpoint_url = definition.request.endpoint_url
        if endpoint_url is None:
            raise ConnectorConfigurationError(
                f"Connector `{definition.connector_key}` does not define a MyEListing start URL."
            )
        start_urls = [endpoint_url]

    seen_listing_keys: set[str] = set()
    seen_sitemap_urls: set[str] = set()
    collected_offers: list[dict[str, Any]] = []

    for start_url in start_urls:
        normalized_start_url = _strip_or_none(start_url)
        if normalized_start_url is None:
            continue
        for offer in _discover_myelisting_offers_from_start_url(
            normalized_start_url,
            definition=definition,
            requested_status=requested_status,
            seen_sitemap_urls=seen_sitemap_urls,
        ):
            detail_url = _strip_or_none(offer.get("url"))
            if detail_url is None:
                continue
            listing_key = (
                _extract_myelisting_listing_key_from_url(detail_url)
                or _strip_or_none(offer.get("source_listing_key"))
            )
            if listing_key is None or listing_key in seen_listing_keys:
                continue
            seen_listing_keys.add(listing_key)
            collected_offers.append(dict(offer))

    return collected_offers


def _discover_myelisting_offers_from_start_url(
    start_url: str,
    *,
    definition: SourceConnectorDefinition,
    requested_status: str | None,
    seen_sitemap_urls: set[str],
) -> list[dict[str, Any]]:
    normalized_start_url = start_url.strip()
    split_result = urlsplit(normalized_start_url)
    normalized_path = split_result.path.lower()

    if normalized_path.endswith(".xml"):
        return _extract_myelisting_offers_from_sitemap_xml(
            normalized_start_url,
            definition=definition,
            seen_sitemap_urls=seen_sitemap_urls,
            url_pattern=definition.request.record_pattern,
        )
    if "/sitemap/" in normalized_path:
        sitemap_html = _download_text_from_url(
            normalized_start_url,
            definition=definition,
            encoding=definition.request.text_encoding,
        )
        search_urls = _extract_myelisting_search_urls_from_state_sitemap_html(
            sitemap_html,
            requested_status=requested_status,
        )
        offers: list[dict[str, Any]] = []
        for search_url in search_urls:
            offers.extend(
                _collect_myelisting_offer_stubs_from_search_pages(
                    search_url,
                    definition=definition,
                )
            )
        return offers
    return _collect_myelisting_offer_stubs_from_search_pages(
        normalized_start_url,
        definition=definition,
    )


def _collect_myelisting_offer_stubs_from_search_pages(
    start_url: str,
    *,
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    parallel_requests = max(1, definition.request.pagination.parallel_requests)
    if parallel_requests > 1:
        return _collect_myelisting_offer_stubs_from_search_pages_parallel(
            start_url,
            definition=definition,
            parallel_requests=parallel_requests,
        )

    return _collect_myelisting_offer_stubs_from_search_pages_sequential(
        start_url,
        definition=definition,
    )


def _collect_myelisting_offer_stubs_from_search_pages_sequential(
    start_url: str,
    *,
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    max_pages = definition.request.pagination.max_pages
    collected_offers: list[dict[str, Any]] = []
    page_number = 1

    while True:
        offers = _fetch_myelisting_search_page_offer_stubs(
            start_url,
            page_number,
            definition=definition,
        )
        if not offers:
            break

        collected_offers.extend(offers)
        page_number += 1
        if max_pages is not None and page_number > max_pages:
            break

    return collected_offers


def _collect_myelisting_offer_stubs_from_search_pages_parallel(
    start_url: str,
    *,
    definition: SourceConnectorDefinition,
    parallel_requests: int,
) -> list[dict[str, Any]]:
    max_pages = definition.request.pagination.max_pages
    first_page_offers = _fetch_myelisting_search_page_offer_stubs(
        start_url,
        1,
        definition=definition,
    )
    if not first_page_offers:
        return []
    if max_pages == 1:
        return first_page_offers

    collected_offers = list(first_page_offers)
    next_page_number = 2
    while True:
        page_numbers = list(range(next_page_number, next_page_number + parallel_requests))
        if max_pages is not None:
            page_numbers = [
                page_number for page_number in page_numbers if page_number <= max_pages
            ]
        if not page_numbers:
            break

        page_results: dict[int, list[dict[str, Any]]] = {}
        with ThreadPoolExecutor(max_workers=parallel_requests) as executor:
            future_to_page_number = {
                executor.submit(
                    _fetch_myelisting_search_page_offer_stubs,
                    start_url,
                    page_number,
                    definition=definition,
                ): page_number
                for page_number in page_numbers
            }
            for future in as_completed(future_to_page_number):
                page_results[future_to_page_number[future]] = future.result()

        should_stop = False
        for page_number in page_numbers:
            offers = page_results.get(page_number, [])
            if not offers:
                should_stop = True
                break
            collected_offers.extend(offers)

        if should_stop:
            break
        next_page_number += len(page_numbers)

    return collected_offers


def _fetch_myelisting_search_page_offer_stubs(
    start_url: str,
    page_number: int,
    *,
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    page_url = _build_myelisting_page_url(start_url, page_number)
    try:
        search_html = _download_text_from_url(
            page_url,
            definition=definition,
            encoding=definition.request.text_encoding,
        )
    except HTTPError as exc:
        if exc.code == 404 and page_number > 1:
            return []
        raise

    try:
        offers = _extract_myelisting_search_offers(search_html)
    except ConnectorExecutionError as exc:
        if page_number > 1 and "pageListings" in str(exc):
            return []
        raise
    normalized_offers: list[dict[str, Any]] = []
    for offer in offers:
        detail_url = _strip_or_none(offer.get("url"))
        if detail_url is None:
            continue
        normalized_offer = dict(offer)
        normalized_offer["page_url"] = page_url
        normalized_offers.append(normalized_offer)
    return normalized_offers


def _extract_myelisting_offers_from_sitemap_xml(
    sitemap_url: str,
    *,
    definition: SourceConnectorDefinition,
    seen_sitemap_urls: set[str],
    url_pattern: str | None,
) -> list[dict[str, Any]]:
    if sitemap_url in seen_sitemap_urls:
        return []
    seen_sitemap_urls.add(sitemap_url)

    sitemap_xml = _download_text_from_url(
        sitemap_url,
        definition=definition,
        encoding=definition.request.text_encoding,
    )
    root = ElementTree.fromstring(sitemap_xml.lstrip())
    root_tag = _strip_xml_namespace(root.tag)

    if root_tag == "sitemapindex":
        offers: list[dict[str, Any]] = []
        for loc_node in root.findall("./{*}sitemap/{*}loc"):
            child_sitemap_url = _strip_or_none(loc_node.text)
            if child_sitemap_url is None or "sitemap_listings" not in child_sitemap_url:
                continue
            offers.extend(
                _extract_myelisting_offers_from_sitemap_xml(
                    child_sitemap_url,
                    definition=definition,
                    seen_sitemap_urls=seen_sitemap_urls,
                    url_pattern=url_pattern,
                )
            )
        return offers

    if root_tag != "urlset":
        raise ConnectorExecutionError(
            f"MyEListing sitemap `{sitemap_url}` did not resolve to a supported XML sitemap type."
        )

    offers: list[dict[str, Any]] = []
    for loc_node in root.findall("./{*}url/{*}loc"):
        detail_url = _strip_or_none(loc_node.text)
        if detail_url is None or "/listing/" not in detail_url:
            continue
        if url_pattern is not None and re.search(url_pattern, detail_url, re.IGNORECASE) is None:
            continue
        offers.append(
            _build_myelisting_offer_stub(
                detail_url,
                page_url=sitemap_url,
            )
        )
    return offers


def _extract_myelisting_search_urls_from_state_sitemap_html(
    sitemap_html: str,
    *,
    requested_status: str | None,
) -> list[str]:
    pattern = re.compile(
        r'href=["\'](?P<url>https://myelisting\.com/properties/for-(?:sale|lease)/[^"\']+/all-property-types/)["\']',
        re.IGNORECASE,
    )
    desired_path_fragment = (
        f"/for-{requested_status}/" if requested_status in {"sale", "lease"} else None
    )
    search_urls: list[str] = []
    seen_urls: set[str] = set()
    for match in pattern.finditer(sitemap_html):
        search_url = match.group("url").strip()
        if desired_path_fragment is not None and desired_path_fragment not in search_url:
            continue
        if _is_myelisting_statewide_search_url(search_url):
            continue
        if search_url in seen_urls:
            continue
        seen_urls.add(search_url)
        search_urls.append(search_url)
    return search_urls


def _is_myelisting_statewide_search_url(url: str) -> bool:
    path_parts = [part for part in urlsplit(url).path.split("/") if part]
    if len(path_parts) < 4:
        return False
    location_slug = path_parts[2]
    return "-" not in location_slug


def _build_myelisting_offer_stub(
    detail_url: str,
    *,
    page_url: str,
) -> dict[str, Any]:
    return {
        "source_listing_key": _extract_myelisting_listing_key_from_url(detail_url),
        "name": _derive_myelisting_listing_title_from_url(detail_url),
        "url": detail_url,
        "price": None,
        "page_url": page_url,
    }


def _build_myelisting_page_url(base_url: str, page_number: int) -> str:
    normalized_base_url = base_url.strip()
    if page_number <= 1:
        return normalized_base_url

    split_result = urlsplit(normalized_base_url)
    normalized_path = re.sub(r"/page-\d+/?$", "/", split_result.path)
    if not normalized_path.endswith("/"):
        normalized_path = f"{normalized_path}/"
    paged_path = f"{normalized_path}page-{page_number}/"
    return urlunsplit(
        (
            split_result.scheme,
            split_result.netloc,
            paged_path,
            split_result.query,
            split_result.fragment,
        )
    )


def _extract_myelisting_search_offers(search_html: str) -> list[dict[str, Any]]:
    script_payload = _extract_html_script_by_id(search_html, "pageListings")
    page_payload = json.loads(script_payload)
    about_items = page_payload.get("about")
    if not isinstance(about_items, list):
        raise ConnectorExecutionError(
            "MyEListing search results did not contain a valid `about` listing array."
        )

    offers: list[dict[str, Any]] = []
    for about_item in about_items:
        if not isinstance(about_item, dict):
            continue
        offer = about_item.get("item")
        if not isinstance(offer, dict):
            continue
        offers.append(
            {
                "source_listing_key": _extract_myelisting_listing_key_from_url(
                    _strip_or_none(offer.get("url"))
                ),
                "name": _strip_or_none(offer.get("name")),
                "url": _strip_or_none(offer.get("url")),
                "price": _strip_or_none(offer.get("price")),
            }
        )
    return offers


def _extract_myelisting_detail_record(
    detail_html: str,
    *,
    detail_url: str,
    offer: dict[str, Any],
    page_url: str,
) -> dict[str, Any]:
    try:
        payload = json.loads(_extract_javascript_variable_value(detail_html, "pdfdata"))
    except (ConnectorExecutionError, json.JSONDecodeError):
        return _build_myelisting_fallback_record(
            detail_url=detail_url,
            offer=offer,
            page_url=page_url,
        )

    listing = payload.get("listing")
    if not isinstance(listing, dict):
        return _build_myelisting_fallback_record(
            detail_url=detail_url,
            offer=offer,
            page_url=page_url,
        )

    source_listing_key = _strip_or_none(
        listing.get("ID")
    ) or _extract_myelisting_listing_key_from_url(detail_url)
    if source_listing_key is None:
        raise ConnectorExecutionError("MyEListing detail page is missing a listing identifier.")

    listing_title = _strip_or_none(listing.get("listing_title")) or _strip_or_none(
        offer.get("name")
    )
    if listing_title is None:
        raise ConnectorExecutionError(
            f"MyEListing listing `{source_listing_key}` is missing a listing title."
        )

    broker_name = _extract_myelisting_broker_name(payload.get("agents"))
    return {
        "source_listing_key": source_listing_key,
        "listing_title": listing_title,
        "listing_status": _normalize_myelisting_listing_status(
            listing_type=listing.get("listing_type"),
            listing_islease=listing.get("listing_islease"),
            page_url=page_url,
        ),
        "asset_type": _strip_or_none(listing.get("proptype")),
        "asking_price": _normalize_myelisting_price(
            call_price=listing.get("call_price"),
            price=listing.get("price"),
        ),
        "acreage": _normalize_positive_numeric_string(listing.get("lot_acre")),
        "building_sqft": _normalize_positive_numeric_string(listing.get("build_sf")),
        "address_line1": _strip_or_none(listing.get("listing_address")),
        "city": _strip_or_none(listing.get("listing_city")),
        "state_code": _strip_or_none(listing.get("listing_state")),
        "postal_code": _strip_or_none(listing.get("listing_zip")),
        "latitude": _normalize_decimal_string(listing.get("listing_lat")),
        "longitude": _normalize_decimal_string(listing.get("listing_lng")),
        "broker_name": broker_name,
        "source_url": detail_url,
        "listing_refreshed": _normalize_myelisting_timestamp(
            listing.get("listing_refreshed") or listing.get("listing_added")
        ),
    }


def _build_myelisting_fallback_record(
    *,
    detail_url: str,
    offer: dict[str, Any],
    page_url: str,
) -> dict[str, Any]:
    source_listing_key = _extract_myelisting_listing_key_from_url(detail_url)
    if source_listing_key is None:
        raise ConnectorExecutionError("MyEListing offer URL does not contain a listing identifier.")

    asking_price = _normalize_myelisting_price(
        call_price=False,
        price=offer.get("price"),
    )
    return {
        "source_listing_key": source_listing_key,
        "listing_title": _strip_or_none(offer.get("name")) or source_listing_key,
        "listing_status": _normalize_myelisting_listing_status(
            listing_type=None,
            listing_islease=None,
            page_url=page_url,
        ),
        "asset_type": None,
        "asking_price": asking_price,
        "acreage": None,
        "building_sqft": None,
        "address_line1": None,
        "city": None,
        "state_code": "TX",
        "postal_code": None,
        "latitude": None,
        "longitude": None,
        "broker_name": None,
        "source_url": detail_url,
        "listing_refreshed": None,
    }


def _fetch_http_csv_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            payload_bytes = _download_connector_bytes(definition, checkpoint)
            raw_records = _parse_csv_bytes(
                payload_bytes,
                delimiter=definition.request.csv_delimiter,
                encoding=definition.request.csv_encoding,
            )
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError, zipfile.BadZipFile) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_http_xlsx_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            payload_bytes = _download_connector_bytes(definition, checkpoint)
            raw_records = _parse_xlsx_bytes(payload_bytes, definition=definition)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_http_zip_csv_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            payload_bytes = _download_connector_bytes(definition, checkpoint)
            archive_member = _extract_zip_member_bytes(
                payload_bytes,
                preferred_name=definition.request.zip_member_name,
                allowed_suffixes=(".csv", ".txt"),
            )
            raw_records = _parse_csv_bytes(
                archive_member,
                delimiter=definition.request.csv_delimiter,
                encoding=definition.request.csv_encoding,
            )
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError, zipfile.BadZipFile) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_http_zip_shapefile_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            payload_bytes = _download_connector_bytes(definition, checkpoint)
            raw_records = _parse_shapefile_zip_bytes(payload_bytes)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError, zipfile.BadZipFile) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_manual_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` is a manual reference source and "
        "requires out-of-band acquisition before ingestion."
    )


def _fetch_arcgis_feature_service_records(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> ConnectorFetchResult:
    last_exception: Exception | None = None

    for attempt_index in range(1, definition.fetch_policy.max_attempts + 1):
        try:
            _RATE_LIMITER.wait(
                definition.connector_key,
                definition.fetch_policy.rate_limit_per_minute,
            )
            raw_records = _fetch_arcgis_feature_pages(definition, checkpoint)
            filtered_raw_records = _filter_incremental_records(
                raw_records,
                checkpoint,
                definition.fetch_policy.checkpoint_field,
            )
            filtered_raw_records = _apply_record_filters(filtered_raw_records, definition)
            mapped_records = _finalize_records_for_pipeline(filtered_raw_records, definition)
            checkpoint_ts = _max_checkpoint_ts(
                filtered_raw_records,
                definition.fetch_policy.checkpoint_field,
            )
            return ConnectorFetchResult(
                records=mapped_records,
                checkpoint_ts=checkpoint_ts or (checkpoint.checkpoint_ts if checkpoint else None),
                checkpoint_cursor=None,
                attempt_count=attempt_index,
            )
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_exception = exc
            if attempt_index < definition.fetch_policy.max_attempts:
                time.sleep(definition.fetch_policy.backoff_seconds)

    raise ConnectorExecutionError(
        f"Connector `{definition.connector_key}` failed after "
        f"{definition.fetch_policy.max_attempts} attempt(s): {last_exception}"
    ) from last_exception


def _fetch_arcgis_feature_pages(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> list[dict[str, Any]]:
    pagination = definition.request.pagination
    page_size = pagination.page_size
    max_pages = pagination.max_pages

    if pagination.strategy not in {"none", "arcgis_offset"}:
        raise ConnectorConfigurationError(
            f"Unsupported pagination strategy `{pagination.strategy}` for "
            f"connector `{definition.connector_key}`."
        )

    parallel_requests = _arcgis_parallel_requests(definition)
    if parallel_requests <= 1 or max_pages == 1:
        return _fetch_arcgis_feature_pages_sequential(definition, checkpoint)

    first_page = _fetch_arcgis_feature_page(
        definition,
        checkpoint,
        offset=0,
        page_size=page_size,
    )
    collected_records = list(first_page.records)
    if not first_page.records or not first_page.has_more:
        return collected_records

    page_index = 1
    step_size = max(len(first_page.records), 1)
    next_offset = step_size
    while max_pages is None or page_index < max_pages:
        remaining_page_slots = (
            parallel_requests
            if max_pages is None
            else min(parallel_requests, max_pages - page_index)
        )
        offsets = [
            next_offset + step_size * offset_index
            for offset_index in range(remaining_page_slots)
        ]
        pages_by_offset = _fetch_arcgis_feature_page_batch(
            definition,
            checkpoint,
            offsets=offsets,
            page_size=page_size,
        )

        should_stop = False
        for offset in offsets:
            page = pages_by_offset[offset]
            if not page.records:
                should_stop = True
                break

            collected_records.extend(page.records)
            page_index += 1
            if max_pages is not None and page_index >= max_pages:
                should_stop = True
                break
            if not page.has_more:
                should_stop = True
                break

        if should_stop:
            break
        next_offset += step_size * len(offsets)

    return collected_records


@dataclass(slots=True)
class _ArcgisFeaturePage:
    records: list[dict[str, Any]]
    has_more: bool


def _fetch_arcgis_feature_pages_sequential(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> list[dict[str, Any]]:
    pagination = definition.request.pagination
    page_size = pagination.page_size
    max_pages = pagination.max_pages
    page_index = 0
    offset = 0
    collected_records: list[dict[str, Any]] = []

    while True:
        page = _fetch_arcgis_feature_page(
            definition,
            checkpoint,
            offset=offset,
            page_size=page_size,
        )
        if not page.records:
            break

        collected_records.extend(page.records)
        page_index += 1
        if max_pages is not None and page_index >= max_pages:
            break
        if not page.has_more:
            break
        offset += len(page.records)

    return collected_records


def _fetch_arcgis_feature_page_batch(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
    *,
    offsets: list[int],
    page_size: int,
) -> dict[int, _ArcgisFeaturePage]:
    pages_by_offset: dict[int, _ArcgisFeaturePage] = {}
    max_workers = min(len(offsets), _arcgis_parallel_requests(definition))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(
                _fetch_arcgis_feature_page,
                definition,
                checkpoint,
                offset=offset,
                page_size=page_size,
            ): offset
            for offset in offsets
        }
        for future in as_completed(futures):
            pages_by_offset[futures[future]] = future.result()
    return pages_by_offset


def _fetch_arcgis_feature_page(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
    *,
    offset: int,
    page_size: int,
) -> _ArcgisFeaturePage:
    pagination = definition.request.pagination
    extra_query_params: dict[str, str] = {
        pagination.offset_param: str(offset),
        pagination.page_size_param: str(page_size),
    }
    request = _build_request(
        definition,
        checkpoint,
        extra_query_params=extra_query_params,
        force_query_path=True,
    )
    with _open_url(request, timeout=definition.request.timeout_seconds) as response:
        payload = json.loads(response.read().decode("utf-8"))

    page_records = _extract_arcgis_feature_list(payload)
    return _ArcgisFeaturePage(
        records=page_records,
        has_more=_arcgis_payload_has_more(
            payload,
            page_size=page_size,
            record_count=len(page_records),
        ),
    )


def _arcgis_payload_has_more(
    payload: Any,
    *,
    page_size: int,
    record_count: int,
) -> bool:
    if not record_count:
        return False
    if isinstance(payload, dict) and payload.get("exceededTransferLimit") is True:
        return True
    return record_count >= page_size


def _arcgis_parallel_requests(definition: SourceConnectorDefinition) -> int:
    requested = definition.request.pagination.parallel_requests
    return max(1, min(requested, 16))


def _download_connector_bytes(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
) -> bytes:
    _RATE_LIMITER.wait(
        definition.connector_key,
        definition.fetch_policy.rate_limit_per_minute,
    )
    request = _build_request(definition, checkpoint)
    with _open_url(request, timeout=definition.request.timeout_seconds) as response:
        return response.read()


def _parse_csv_bytes(
    payload_bytes: bytes,
    *,
    delimiter: str,
    encoding: str,
) -> list[dict[str, Any]]:
    decoded = payload_bytes.decode(encoding)
    reader = csv.DictReader(StringIO(decoded), delimiter=delimiter)
    return [dict(row) for row in reader]


def _parse_xlsx_bytes(
    payload_bytes: bytes,
    *,
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(payload_bytes), read_only=True, data_only=True)
    try:
        try:
            if definition.request.xlsx_sheet_name is not None:
                worksheet = workbook[definition.request.xlsx_sheet_name]
            else:
                worksheet = workbook[workbook.sheetnames[0]]
        except KeyError as exc:
            raise ValueError(
                f"Workbook sheet `{definition.request.xlsx_sheet_name}` was not found."
            ) from exc

        header_row_number = definition.request.xlsx_header_row
        data_start_row = definition.request.xlsx_data_start_row or (header_row_number + 1)
        header_values = next(
            worksheet.iter_rows(
                min_row=header_row_number,
                max_row=header_row_number,
                values_only=True,
            ),
            None,
        )
        if header_values is None:
            return []

        headers = [
            _normalize_optional_string(value) or f"column_{index}"
            for index, value in enumerate(header_values, start=1)
        ]
        raw_records: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
            if row is None or all(value is None or str(value).strip() == "" for value in row):
                continue
            raw_records.append(
                {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers)
                }
            )
        return raw_records
    finally:
        workbook.close()


def _parse_html_records(
    payload_bytes: bytes,
    *,
    encoding: str,
    record_pattern: str | None,
) -> list[dict[str, Any]]:
    if record_pattern is None:
        raise ConnectorConfigurationError(
            "HTML connector requests require `request.record_pattern`."
        )

    decoded = payload_bytes.decode(encoding)
    pattern = re.compile(record_pattern, re.IGNORECASE | re.DOTALL)
    records: list[dict[str, Any]] = []
    for match in pattern.finditer(decoded):
        records.append(
            {
                key: _clean_html_text(value)
                for key, value in match.groupdict().items()
            }
        )
    return records


def _parse_first_html_table_records(source_text: str) -> list[dict[str, Any]]:
    table_match = re.search(
        r"<table[^>]*>(?P<table>.*?)</table>",
        source_text,
        re.IGNORECASE | re.DOTALL,
    )
    if table_match is None:
        return []

    row_payloads = re.findall(
        r"<tr[^>]*>(?P<row>.*?)</tr>",
        table_match.group("table"),
        re.IGNORECASE | re.DOTALL,
    )
    cleaned_rows: list[list[str | None]] = []
    for row_payload in row_payloads:
        cells = [
            _clean_html_text(cell_payload)
            for cell_payload in re.findall(
                r"<t[hd][^>]*>(.*?)</t[hd]>",
                row_payload,
                re.IGNORECASE | re.DOTALL,
            )
        ]
        if any(cell is not None for cell in cells):
            cleaned_rows.append(cells)

    if len(cleaned_rows) < 2:
        return []

    headers = [
        header or f"column_{index}"
        for index, header in enumerate(cleaned_rows[0], start=1)
    ]
    records: list[dict[str, Any]] = []
    for row in cleaned_rows[1:]:
        normalized_row = list(row)
        if len(normalized_row) < len(headers):
            normalized_row.extend([None] * (len(headers) - len(normalized_row)))
        records.append(
            {
                header: normalized_row[index]
                for index, header in enumerate(headers)
            }
        )
    return records


def _download_text_from_url(
    url: str,
    *,
    definition: SourceConnectorDefinition,
    encoding: str,
) -> str:
    return _download_url_bytes(url, definition=definition).decode(encoding)


def _download_url_bytes(
    url: str,
    *,
    definition: SourceConnectorDefinition,
) -> bytes:
    _RATE_LIMITER.wait(
        definition.connector_key,
        definition.fetch_policy.rate_limit_per_minute,
    )
    request = Request(
        url=url,
        headers=dict(definition.request.headers),
        method="GET",
    )
    with _open_url(request, timeout=definition.request.timeout_seconds) as response:
        return response.read()


def _append_query_params(url: str, query_params: dict[str, Any]) -> str:
    split_url = urlsplit(url)
    existing_params: dict[str, list[str]] = {}
    for component in split_url.query.split("&"):
        if "=" not in component:
            continue
        key, value = component.split("=", 1)
        existing_params.setdefault(key, []).append(value)

    for key, value in query_params.items():
        normalized = _normalize_optional_string(value)
        if normalized is None:
            continue
        existing_params[key] = [normalized]

    flat_params = [
        (key, value)
        for key, values in existing_params.items()
        for value in values
    ]
    return urlunsplit(
        (
            split_url.scheme,
            split_url.netloc,
            split_url.path,
            urlencode(flat_params),
            split_url.fragment,
        )
    )


def _load_local_geojson_geometry(boundary_geojson_path: str):
    boundary_path = Path(boundary_geojson_path)
    if not boundary_path.is_absolute():
        boundary_path = boundary_path.resolve()
    if not boundary_path.exists():
        raise FileNotFoundError(f"Boundary GeoJSON path `{boundary_path}` does not exist.")

    payload = json.loads(boundary_path.read_text(encoding="utf-8"))
    if payload.get("type") == "FeatureCollection":
        geometry_payload = {
            "type": "GeometryCollection",
            "geometries": [
                feature.get("geometry")
                for feature in payload.get("features", [])
                if isinstance(feature, dict) and feature.get("geometry") is not None
            ],
        }
    elif payload.get("type") == "Feature":
        geometry_payload = payload.get("geometry")
    else:
        geometry_payload = payload

    try:
        geometry = shapely_shape(geometry_payload)
    except (AttributeError, TypeError, ValueError, ShapelyError) as exc:
        raise ConnectorExecutionError("Boundary GeoJSON is invalid.") from exc
    if geometry.is_empty:
        raise ConnectorExecutionError("Boundary GeoJSON cannot resolve to an empty geometry.")
    return geometry


def _iter_grid_points_within_geometry(
    geometry,
    grid_step_degrees: float,
) -> list[tuple[float, float]]:
    min_x, min_y, max_x, max_y = geometry.bounds
    longitude = min_x + (grid_step_degrees / 2)
    candidate_points: list[tuple[float, float]] = []
    while longitude <= max_x:
        latitude = min_y + (grid_step_degrees / 2)
        while latitude <= max_y:
            point = Point(longitude, latitude)
            if geometry.covers(point):
                candidate_points.append((round(latitude, 6), round(longitude, 6)))
            latitude += grid_step_degrees
        longitude += grid_step_degrees
    return candidate_points


def _extract_html_script_by_id(source_text: str, script_id: str) -> str:
    pattern = re.compile(
        rf'<script[^>]*id=["\']{re.escape(script_id)}["\'][^>]*>(?P<payload>.*?)</script>',
        re.IGNORECASE | re.DOTALL,
    )
    match = pattern.search(source_text)
    if match is None:
        raise ConnectorExecutionError(
            f"Unable to locate HTML script payload `{script_id}` in connector response."
        )
    return match.group("payload").strip()


def _extract_javascript_variable_value(source_text: str, variable_name: str) -> str:
    match = re.search(
        rf"\bvar\s+{re.escape(variable_name)}\s*=\s*",
        source_text,
        re.IGNORECASE,
    )
    if match is None:
        raise ConnectorExecutionError(
            f"Unable to locate JavaScript variable `{variable_name}` in connector response."
        )

    start_index = match.end()
    while start_index < len(source_text) and source_text[start_index].isspace():
        start_index += 1

    if start_index >= len(source_text) or source_text[start_index] not in "{[":
        raise ConnectorExecutionError(
            f"JavaScript variable `{variable_name}` does not contain an object payload."
        )

    opening_character = source_text[start_index]
    closing_character = "}" if opening_character == "{" else "]"
    nesting_depth = 0
    active_quote: str | None = None
    escaped = False

    for cursor in range(start_index, len(source_text)):
        character = source_text[cursor]
        if active_quote is not None:
            if escaped:
                escaped = False
                continue
            if character == "\\":
                escaped = True
                continue
            if character == active_quote:
                active_quote = None
            continue

        if character in {'"', "'"}:
            active_quote = character
            continue
        if character == opening_character:
            nesting_depth += 1
            continue
        if character == closing_character:
            nesting_depth -= 1
            if nesting_depth == 0:
                return source_text[start_index : cursor + 1]

    raise ConnectorExecutionError(
        f"JavaScript variable `{variable_name}` did not terminate cleanly."
    )


def _extract_zip_member_bytes(
    payload_bytes: bytes,
    *,
    preferred_name: str | None,
    allowed_suffixes: tuple[str, ...],
) -> bytes:
    with zipfile.ZipFile(BytesIO(payload_bytes)) as archive:
        names = archive.namelist()
        if preferred_name is not None:
            for name in names:
                if name.lower().endswith(preferred_name.lower()):
                    return archive.read(name)
            raise ConnectorExecutionError(
                f"Zip archive does not contain requested member `{preferred_name}`."
            )

        for name in names:
            lowered = name.lower()
            if lowered.endswith(allowed_suffixes):
                return archive.read(name)

    raise ConnectorExecutionError("Zip archive did not contain a supported member.")


def _parse_shapefile_zip_bytes(payload_bytes: bytes) -> list[dict[str, Any]]:
    with zipfile.ZipFile(BytesIO(payload_bytes)) as archive:
        members = {name.lower(): name for name in archive.namelist()}
        shp_name = next((name for name in members if name.endswith(".shp")), None)
        shx_name = next((name for name in members if name.endswith(".shx")), None)
        dbf_name = next((name for name in members if name.endswith(".dbf")), None)
        prj_name = next((name for name in members if name.endswith(".prj")), None)
        if shp_name is None or shx_name is None or dbf_name is None:
            raise ConnectorExecutionError(
                "Shapefile archive must contain .shp, .shx, and .dbf members."
            )
        geometry_transformer = _build_shapefile_geometry_transformer(
            archive.read(members[prj_name]) if prj_name is not None else None
        )

        reader = shapefile.Reader(
            shp=BytesIO(archive.read(members[shp_name])),
            shx=BytesIO(archive.read(members[shx_name])),
            dbf=BytesIO(archive.read(members[dbf_name])),
        )

        field_names = [field[0] for field in reader.fields if field[0] != "DeletionFlag"]
        records: list[dict[str, Any]] = []
        for shape_record in reader.iterShapeRecords():
            if shape_record.shape.shapeType == shapefile.NULL:
                continue
            record = dict(zip(field_names, shape_record.record, strict=False))
            native_geometry = shapely_shape(shape_record.shape.__geo_interface__)
            if native_geometry.is_empty:
                continue
            record["__geometry_native_area__"] = native_geometry.area
            geometry = native_geometry.__geo_interface__
            if geometry_transformer is not None:
                geometry = shapely_transform(
                    geometry_transformer,
                    native_geometry,
                ).__geo_interface__
            record["__geometry__"] = geometry
            records.append(record)
        return records


def _build_shapefile_geometry_transformer(prj_bytes: bytes | None):
    if prj_bytes is None:
        return None

    try:
        prj_text = prj_bytes.decode("utf-8")
    except UnicodeDecodeError:
        prj_text = prj_bytes.decode("latin-1")

    try:
        source_crs = CRS.from_wkt(prj_text)
    except Exception as exc:  # pragma: no cover - pyproj error surface varies by version
        raise ConnectorExecutionError("Unable to parse shapefile projection metadata.") from exc

    target_crs = CRS.from_epsg(4326)
    if source_crs == target_crs:
        return None

    transformer = Transformer.from_crs(source_crs, target_crs, always_xy=True)
    return transformer.transform


def _normalize_arcgis_geometry(
    geometry: Any,
    *,
    spatial_reference: Any,
) -> tuple[dict[str, Any] | None, float | None]:
    if not isinstance(geometry, dict):
        return None, None
    if _is_geojson_geometry(geometry):
        return geometry, None

    try:
        native_geometry = _arcgis_geometry_to_shapely(geometry)
    except ConnectorExecutionError:
        return None, None

    native_area = native_geometry.area
    geographic_geometry = native_geometry
    transformer = _build_arcgis_geometry_transformer(spatial_reference)
    if transformer is not None:
        geographic_geometry = shapely_transform(transformer, native_geometry)

    return geographic_geometry.__geo_interface__, native_area


def _build_arcgis_geometry_transformer(spatial_reference: Any):
    source_wkid = _extract_arcgis_wkid(spatial_reference)
    if source_wkid is None:
        return None

    return _build_arcgis_geometry_transformer_for_wkid(source_wkid)


@lru_cache(maxsize=32)
def _build_arcgis_geometry_transformer_for_wkid(source_wkid: int):
    source_crs = CRS.from_epsg(source_wkid)
    target_crs = CRS.from_epsg(4326)
    if source_crs == target_crs:
        return None

    transformer = Transformer.from_crs(source_crs, target_crs, always_xy=True)
    return transformer.transform


def _is_geojson_geometry(geometry: dict[str, Any]) -> bool:
    return isinstance(geometry.get("type"), str) and (
        "coordinates" in geometry or "geometries" in geometry
    )


def _extract_arcgis_wkid(spatial_reference: Any) -> int | None:
    if isinstance(spatial_reference, int):
        return spatial_reference
    if isinstance(spatial_reference, str) and spatial_reference.strip().isdigit():
        return int(spatial_reference.strip())
    if not isinstance(spatial_reference, dict):
        return None

    for key in ("latestWkid", "wkid"):
        candidate = spatial_reference.get(key)
        if isinstance(candidate, int):
            return candidate
        if isinstance(candidate, str) and candidate.strip().isdigit():
            return int(candidate.strip())
    return None


def _arcgis_geometry_to_shapely(geometry: dict[str, Any]):
    if "x" in geometry and "y" in geometry:
        coordinate = _sanitize_arcgis_coordinate([geometry["x"], geometry["y"]])
        if coordinate is None:
            raise ConnectorExecutionError("ArcGIS point geometry contains invalid coordinates.")
        return Point(coordinate)

    paths = geometry.get("paths")
    if isinstance(paths, list):
        line_strings = [
            LineString(coordinates)
            for path in paths
            if (coordinates := _sanitize_arcgis_coordinate_sequence(path, close_ring=False))
            and len(coordinates) >= 2
        ]
        if not line_strings:
            raise ConnectorExecutionError("ArcGIS geometry contains no path coordinates.")
        if len(line_strings) == 1:
            return line_strings[0]
        return MultiLineString(line_strings)

    rings = geometry.get("rings")
    if isinstance(rings, list):
        polygons = _arcgis_rings_to_polygons(rings)
        if not polygons:
            raise ConnectorExecutionError("ArcGIS geometry contains no polygon coordinates.")
        if len(polygons) == 1:
            return polygons[0]
        return MultiPolygon(polygons)

    points = geometry.get("points")
    if isinstance(points, list):
        point_geometries = [
            Point(coordinate)
            for point in points
            if (coordinate := _sanitize_arcgis_coordinate(point)) is not None
        ]
        if not point_geometries:
            raise ConnectorExecutionError("ArcGIS geometry contains no point coordinates.")
        if len(point_geometries) == 1:
            return point_geometries[0]
        return MultiPoint(point_geometries)

    raise ConnectorExecutionError("Unsupported ArcGIS geometry payload.")


def _arcgis_rings_to_polygons(rings: list[Any]) -> list[Polygon]:
    polygons: list[Polygon] = []
    shells: list[Polygon] = []

    for ring in rings:
        normalized_ring = _sanitize_arcgis_coordinate_sequence(ring)
        if len(normalized_ring) < 4:
            continue
        ring_polygon = Polygon(normalized_ring)
        if ring_polygon.is_empty:
            continue
        if not ring_polygon.is_valid:
            ring_polygon = ring_polygon.buffer(0)
        if ring_polygon.is_empty:
            continue
        if ring_polygon.geom_type == "Polygon":
            if ring_polygon.exterior.is_ccw:
                shells.append(ring_polygon)
                continue
            _assign_arcgis_hole(polygons, shells, ring_polygon)
            continue
        if ring_polygon.geom_type == "MultiPolygon":
            for component in ring_polygon.geoms:
                if component.exterior.is_ccw:
                    shells.append(component)
                else:
                    _assign_arcgis_hole(polygons, shells, component)

    polygons.extend(shells)
    return polygons


def _sanitize_arcgis_coordinate(point: Any) -> tuple[float, float] | None:
    if not isinstance(point, (list, tuple)) or len(point) < 2:
        return None
    try:
        x = float(point[0])
        y = float(point[1])
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x) or not math.isfinite(y):
        return None
    return (x, y)


def _sanitize_arcgis_coordinate_sequence(
    sequence: Any,
    *,
    close_ring: bool = True,
) -> list[tuple[float, float]]:
    if not isinstance(sequence, list):
        return []

    coordinates = [
        coordinate
        for point in sequence
        if (coordinate := _sanitize_arcgis_coordinate(point)) is not None
    ]
    if close_ring and len(coordinates) >= 3 and coordinates[0] != coordinates[-1]:
        coordinates.append(coordinates[0])
    return coordinates


def _assign_arcgis_hole(
    polygons: list[Polygon],
    shells: list[Polygon],
    hole_polygon: Polygon,
) -> None:
    hole_coordinates = list(hole_polygon.exterior.coords)
    for shell_index in range(len(shells) - 1, -1, -1):
        shell_polygon = shells[shell_index]
        if not shell_polygon.contains(hole_polygon.representative_point()):
            continue
        shells[shell_index] = Polygon(
            shell_polygon.exterior.coords,
            holes=[*shell_polygon.interiors, hole_coordinates],
        )
        return

    for polygon_index in range(len(polygons) - 1, -1, -1):
        polygon = polygons[polygon_index]
        if not polygon.contains(hole_polygon.representative_point()):
            continue
        polygons[polygon_index] = Polygon(
            polygon.exterior.coords,
            holes=[*polygon.interiors, hole_coordinates],
        )
        return

    polygons.append(Polygon(hole_coordinates))


def _finalize_records_for_pipeline(
    raw_records: list[dict[str, Any]],
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    preprocess_strategy = (definition.preprocess_strategy or "").strip().lower()
    if preprocess_strategy in {
        "expand_evidence_attributes",
        "spatial_filter_expand_evidence_attributes",
    }:
        return [dict(record) for record in raw_records]
    return [_map_record(raw_record, definition) for raw_record in raw_records]


def _build_request(
    definition: SourceConnectorDefinition,
    checkpoint: ConnectorCheckpoint | None,
    *,
    extra_query_params: dict[str, str] | None = None,
    force_query_path: bool = False,
) -> Request:
    endpoint_url = definition.request.endpoint_url
    if endpoint_url is None:
        raise ConnectorConfigurationError(
            f"Connector `{definition.connector_key}` does not define an endpoint URL."
        )

    normalized_endpoint = endpoint_url.rstrip("/")
    if force_query_path and not normalized_endpoint.endswith("/query"):
        normalized_endpoint = f"{normalized_endpoint}/query"

    query_params = dict(definition.request.query_params)
    if force_query_path:
        query_params.setdefault("where", "1=1")
        query_params.setdefault("outFields", "*")
        query_params.setdefault("returnGeometry", "true")
        query_params.setdefault("f", "geojson")
    if extra_query_params:
        query_params.update(extra_query_params)
    if (
        checkpoint is not None
        and checkpoint.checkpoint_ts is not None
        and definition.fetch_policy.checkpoint_param is not None
    ):
        query_params[definition.fetch_policy.checkpoint_param] = (
            checkpoint.checkpoint_ts.isoformat()
        )

    encoded_query = urlencode(query_params)
    target_url = (
        normalized_endpoint
        if not encoded_query
        else f"{normalized_endpoint}?{encoded_query}"
    )
    headers = dict(definition.request.headers)
    method = definition.request.method.strip().upper()

    if definition.request.auth_header_name and definition.request.auth_env_var:
        token = os.getenv(definition.request.auth_env_var)
        if not token:
            raise ConnectorConfigurationError(
                f"Connector `{definition.connector_key}` requires env var "
                f"`{definition.request.auth_env_var}`."
            )
        headers[definition.request.auth_header_name] = token

    if definition.request.auth_query_param_name and definition.request.auth_env_var:
        token = os.getenv(definition.request.auth_env_var)
        if not token:
            raise ConnectorConfigurationError(
                f"Connector `{definition.connector_key}` requires env var "
                f"`{definition.request.auth_env_var}`."
            )
        auth_query = urlencode({definition.request.auth_query_param_name: token})
        separator = "&" if "?" in target_url else "?"
        target_url = f"{target_url}{separator}{auth_query}"

    data: bytes | None = None
    if definition.request.json_body:
        data = json.dumps(definition.request.json_body).encode("utf-8")
        headers.setdefault("Content-Type", "application/json")
    elif definition.request.body_text is not None:
        data = definition.request.body_text.encode("utf-8")

    return Request(url=target_url, headers=headers, data=data, method=method)


def _extract_record_list(payload: Any, record_path: list[str]) -> list[dict[str, Any]]:
    extracted = payload
    for path_part in record_path:
        if not isinstance(extracted, dict) or path_part not in extracted:
            raise ConnectorExecutionError(
                f"Unable to resolve record path `{'/'.join(record_path)}` in connector payload."
            )
        extracted = extracted[path_part]

    if isinstance(extracted, list):
        return _normalize_record_list(extracted)
    raise ConnectorExecutionError("Connector payload did not resolve to a list of records.")


def _normalize_record_list(extracted: list[Any]) -> list[dict[str, Any]]:
    if not extracted:
        return []

    if all(isinstance(item, dict) for item in extracted):
        return [dict(item) for item in extracted]

    first_row = extracted[0]
    if (
        isinstance(first_row, list)
        and first_row
        and all(not isinstance(value, (dict, list)) for value in first_row)
    ):
        headers = [str(value) for value in first_row]
        normalized_records: list[dict[str, Any]] = []
        for row in extracted[1:]:
            if not isinstance(row, list):
                raise ConnectorExecutionError(
                    "Connector payload mixes header-based rows with non-list records."
                )
            if len(row) != len(headers):
                raise ConnectorExecutionError(
                    "Connector payload contains a header-based row with the wrong column count."
                )
            normalized_records.append(
                {
                    header: value
                    for header, value in zip(headers, row, strict=True)
                }
            )
        return normalized_records

    raise ConnectorExecutionError(
        "Connector payload did not resolve to a list of record objects."
    )


def _clean_html_text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"<[^>]+>", " ", str(value))
    normalized = html.unescape(normalized)
    normalized = " ".join(normalized.split())
    return normalized or None


def _extract_myelisting_listing_key_from_url(value: Any) -> str | None:
    normalized_url = _strip_or_none(value)
    if normalized_url is None:
        return None
    match = re.search(r"/listing/(?P<listing_id>\d+)/", normalized_url)
    if match is None:
        return None
    return match.group("listing_id")


def _derive_myelisting_listing_title_from_url(value: Any) -> str | None:
    normalized_url = _strip_or_none(value)
    if normalized_url is None:
        return None
    path_parts = [part for part in urlsplit(normalized_url).path.split("/") if part]
    if not path_parts:
        return None
    slug = path_parts[-1]
    if slug.isdigit() and len(path_parts) >= 2:
        slug = path_parts[-2]
    tokens = [token for token in slug.split("-") if token]
    if not tokens:
        return None
    return " ".join(
        token.upper() if len(token) <= 3 and token.isalpha() else token.capitalize()
        for token in tokens
    )


def _derive_acrevalue_listing_title(listing: dict[str, Any]) -> str | None:
    parts = [
        _normalize_positive_numeric_string(listing.get("acres")),
        "Acres" if _normalize_positive_numeric_string(listing.get("acres")) else None,
        _clean_html_text(listing.get("address")),
        _clean_html_text(listing.get("city")),
        _strip_or_none(listing.get("state")),
    ]
    return " ".join(part for part in parts if part is not None) or None


def _normalize_acrevalue_listing_status(value: Any) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    lowered = normalized.lower()
    if "sale" in lowered:
        return "sale"
    if "sold" in lowered:
        return "sold"
    if lowered == "active":
        return "sale"
    return lowered.replace("-", "_").replace(" ", "_")


def _normalize_acrevalue_date(value: Any) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    for date_format in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(normalized, date_format).replace(tzinfo=UTC)
        except ValueError:
            continue
        return parsed.isoformat()
    try:
        return datetime.fromisoformat(normalized.replace("Z", "+00:00")).isoformat()
    except ValueError:
        return None


def _first_clean_acrevalue_listing_value(
    listing: dict[str, Any],
    *field_names: str,
) -> str | None:
    for field_name in field_names:
        value = _clean_html_text(listing.get(field_name))
        if value is not None:
            return value
    return None


def _normalize_public_listing_url(value: str | None, *, base_url: str) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    split_result = urlsplit(normalized)
    if split_result.scheme and split_result.netloc:
        return normalized
    if not normalized.startswith("/"):
        normalized = f"/{normalized}"
    return f"{base_url.rstrip('/')}{normalized}"


def _infer_myelisting_requested_status(
    definition: SourceConnectorDefinition,
) -> str | None:
    candidate_urls = [
        definition.request.endpoint_url,
        *definition.request.start_urls,
    ]
    for candidate_url in candidate_urls:
        normalized_url = _strip_or_none(candidate_url)
        if normalized_url is None:
            continue
        if "/for-sale/" in normalized_url:
            return "sale"
        if "/for-lease/" in normalized_url:
            return "lease"
    return None


def _strip_xml_namespace(tag: str) -> str:
    if "}" not in tag:
        return tag
    return tag.rsplit("}", 1)[-1]


def _normalize_myelisting_listing_status(
    *,
    listing_type: Any,
    listing_islease: Any,
    page_url: str,
) -> str | None:
    normalized_type = _strip_or_none(listing_type)
    if normalized_type is not None:
        return normalized_type.lower().replace("-", "_").replace(" ", "_")

    if str(listing_islease).strip() in {"1", "true", "True"}:
        return "lease"
    if "/for-lease/" in page_url:
        return "lease"
    if "/for-sale/" in page_url:
        return "sale"
    return None


def _normalize_myelisting_price(*, call_price: Any, price: Any) -> str | None:
    if str(call_price).strip() in {"1", "true", "True"}:
        return None
    return _normalize_positive_numeric_string(price)


def _normalize_positive_numeric_string(value: Any) -> str | None:
    normalized = _normalize_decimal_string(value)
    if normalized is None:
        return None
    try:
        numeric_value = float(normalized)
    except ValueError:
        return None
    if numeric_value <= 0:
        return None
    return normalized


def _normalize_decimal_string(value: Any) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    normalized = normalized.replace(",", "")
    try:
        numeric_value = float(normalized)
    except ValueError:
        return None
    if numeric_value.is_integer():
        return str(int(numeric_value))
    return format(numeric_value, "f").rstrip("0").rstrip(".")


def _normalize_myelisting_timestamp(value: Any) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    try:
        parsed = datetime.fromisoformat(normalized.replace(" ", "T"))
    except ValueError:
        return None
    return parsed.isoformat()


def _extract_myelisting_broker_name(agents_payload: Any) -> str | None:
    if not isinstance(agents_payload, list) or not agents_payload:
        return None
    first_agent = agents_payload[0]
    if not isinstance(first_agent, dict):
        return None

    office_name = _strip_or_none(first_agent.get("office_name"))
    full_name = " ".join(
        part
        for part in (
            _strip_or_none(first_agent.get("user_firstname")),
            _strip_or_none(first_agent.get("user_lastname")),
        )
        if part is not None
    )
    if office_name and full_name:
        return f"{office_name} ({full_name})"
    return office_name or full_name or None


def _extract_arcgis_feature_list(payload: Any) -> list[dict[str, Any]]:
    features = payload.get("features") if isinstance(payload, dict) else None
    if not isinstance(features, list):
        raise ConnectorExecutionError("ArcGIS payload does not contain a `features` list.")

    payload_spatial_reference = (
        payload.get("spatialReference") if isinstance(payload, dict) else None
    )
    normalized_records: list[dict[str, Any]] = []
    for feature in features:
        if not isinstance(feature, dict):
            continue
        properties = dict(feature.get("properties") or feature.get("attributes") or {})
        geometry = feature.get("geometry")
        feature_spatial_reference = None
        if isinstance(geometry, dict):
            feature_spatial_reference = geometry.get("spatialReference")
        if feature_spatial_reference is None:
            feature_spatial_reference = payload_spatial_reference

        normalized_geometry, native_area = _normalize_arcgis_geometry(
            geometry,
            spatial_reference=feature_spatial_reference,
        )
        properties["__geometry__"] = normalized_geometry if normalized_geometry is not None else geometry
        if native_area is not None:
            properties["__geometry_native_area__"] = native_area
        if feature.get("id") is not None:
            properties["__feature_id__"] = feature.get("id")
        normalized_records.append(properties)
    return normalized_records


def _filter_incremental_records(
    raw_records: list[dict[str, Any]],
    checkpoint: ConnectorCheckpoint | None,
    checkpoint_field: str | None,
) -> list[dict[str, Any]]:
    if checkpoint is None or checkpoint.checkpoint_ts is None or checkpoint_field is None:
        return [dict(record) for record in raw_records]

    filtered: list[dict[str, Any]] = []
    for raw_record in raw_records:
        raw_value = _resolve_field(raw_record, checkpoint_field)
        checkpoint_ts = _coerce_datetime(raw_value)
        if checkpoint_ts is None or checkpoint_ts > checkpoint.checkpoint_ts:
            filtered.append(dict(raw_record))
    return filtered


def _apply_record_filters(
    raw_records: list[dict[str, Any]],
    definition: SourceConnectorDefinition,
) -> list[dict[str, Any]]:
    if not definition.row_filters:
        return [dict(record) for record in raw_records]

    filtered_records: list[dict[str, Any]] = []
    for raw_record in raw_records:
        if all(_record_matches_filter(raw_record, filter_rule) for filter_rule in definition.row_filters):
            filtered_records.append(dict(raw_record))
    return filtered_records


def _record_matches_filter(
    raw_record: dict[str, Any],
    filter_rule: SourceConnectorRowFilterRule,
) -> bool:
    operator = filter_rule.operator.strip().lower()
    case_sensitive = bool(filter_rule.options.get("case_sensitive", False))
    raw_value = _extract_source_value(raw_record, filter_rule.source)

    if operator == "exists":
        normalized = _normalize_filter_value(raw_value, case_sensitive=case_sensitive)
        return normalized is not None
    if operator == "not_exists":
        normalized = _normalize_filter_value(raw_value, case_sensitive=case_sensitive)
        return normalized is None

    normalized_value = _normalize_filter_value(raw_value, case_sensitive=case_sensitive)
    normalized_expected = _normalize_filter_value(
        filter_rule.value,
        case_sensitive=case_sensitive,
    )
    normalized_expected_values = [
        normalized_candidate
        for normalized_candidate in (
            _normalize_filter_value(candidate, case_sensitive=case_sensitive)
            for candidate in filter_rule.values
        )
        if normalized_candidate is not None
    ]

    if operator == "equals":
        return normalized_value == normalized_expected
    if operator == "not_equals":
        return normalized_value != normalized_expected
    if operator == "startswith":
        return (
            normalized_value is not None
            and normalized_expected is not None
            and normalized_value.startswith(normalized_expected)
        )
    if operator == "endswith":
        return (
            normalized_value is not None
            and normalized_expected is not None
            and normalized_value.endswith(normalized_expected)
        )
    if operator == "in":
        return normalized_value in normalized_expected_values
    if operator == "not_in":
        return normalized_value not in normalized_expected_values
    if operator == "regex":
        pattern = _strip_or_none(filter_rule.value)
        if normalized_value is None or pattern is None:
            return False
        flags = 0 if case_sensitive else re.IGNORECASE
        return re.search(pattern, normalized_value, flags) is not None

    raise ConnectorConfigurationError(
        f"Unsupported record filter operator `{filter_rule.operator}`."
    )


def _normalize_filter_value(value: Any, *, case_sensitive: bool) -> str | None:
    normalized = _strip_or_none(value)
    if normalized is None:
        return None
    return normalized if case_sensitive else normalized.lower()


def _max_checkpoint_ts(
    raw_records: list[dict[str, Any]],
    checkpoint_field: str | None,
) -> datetime | None:
    if checkpoint_field is None:
        return None

    checkpoint_values = [
        checkpoint_ts
        for checkpoint_ts in (
            _coerce_datetime(_resolve_field(record, checkpoint_field)) for record in raw_records
        )
        if checkpoint_ts is not None
    ]
    return max(checkpoint_values, default=None)


def _map_record(
    raw_record: dict[str, Any],
    definition: SourceConnectorDefinition,
) -> dict[str, Any]:
    mapped_record = dict(definition.static_fields)

    if definition.field_rules:
        for field_rule in definition.field_rules:
            mapped_record[field_rule.target] = _apply_field_rule(
                raw_record=raw_record,
                mapped_record=mapped_record,
                field_rule=field_rule,
            )
        return mapped_record

    for target_key, source_key in definition.field_map.items():
        mapped_record[target_key] = _resolve_field(raw_record, source_key)
    return mapped_record


def _apply_field_rule(
    *,
    raw_record: dict[str, Any],
    mapped_record: dict[str, Any],
    field_rule: SourceConnectorFieldRule,
) -> Any:
    context = _SafeFormatDict(
        {
            **raw_record,
            **mapped_record,
        }
    )
    if field_rule.template is not None:
        transformed_value = _apply_transform(
            value=field_rule.template.format_map(context),
            context=context,
            field_rule=field_rule,
        )
    elif isinstance(field_rule.source, list):
        transformed_value = None
        for candidate_source in field_rule.source:
            candidate_value = _resolve_field(raw_record, candidate_source)
            if candidate_value is None or str(candidate_value).strip() == "":
                continue
            transformed_value = _apply_transform(
                value=candidate_value,
                context=_SafeFormatDict(
                    {
                        **context,
                        "__source_name__": candidate_source,
                    }
                ),
                field_rule=field_rule,
            )
            if transformed_value is not None:
                break
    else:
        raw_value = _extract_source_value(raw_record, field_rule.source)
        transformed_value = _apply_transform(
            value=raw_value,
            context=_SafeFormatDict(
                {
                    **context,
                    "__source_name__": field_rule.source,
                }
            ),
            field_rule=field_rule,
        )
    if transformed_value is None and field_rule.default is not None:
        return field_rule.default
    return transformed_value


def resolve_connector_field_rule_value(
    *,
    raw_record: dict[str, Any],
    mapped_record: dict[str, Any],
    field_rule: SourceConnectorFieldRule,
) -> Any:
    return _apply_field_rule(
        raw_record=raw_record,
        mapped_record=mapped_record,
        field_rule=field_rule,
    )


def _apply_transform(
    *,
    value: Any,
    context: dict[str, Any],
    field_rule: SourceConnectorFieldRule,
) -> Any:
    transform = field_rule.transform.lower()
    if transform == "identity":
        return value
    if transform == "stringify":
        return None if value is None else str(value)
    if transform == "strip":
        return _strip_or_none(value)
    if transform == "upper":
        normalized = _strip_or_none(value)
        return normalized.upper() if normalized is not None else None
    if transform == "lower":
        normalized = _strip_or_none(value)
        return normalized.lower() if normalized is not None else None
    if transform == "geojson_to_wkt":
        if value is None:
            return None
        try:
            return shapely_shape(value).wkt
        except (AttributeError, TypeError, ValueError, ShapelyError) as exc:
            if isinstance(value, dict):
                try:
                    return _arcgis_geometry_to_shapely(value).wkt
                except ConnectorExecutionError:
                    pass
            raise ConnectorExecutionError("Unable to convert GeoJSON geometry into WKT.") from exc
    if transform == "square_feet_to_acres":
        if value is None or str(value).strip() == "":
            return None
        acreage_value = float(value) / 43560
        return f"{acreage_value:.6f}".rstrip("0").rstrip(".")
    if transform == "acres_or_square_feet_by_source":
        if value is None or str(value).strip() == "":
            return None
        source_name = _strip_or_none(context.get("__source_name__"))
        if source_name is not None and re.search(r"acre", source_name, re.IGNORECASE):
            if isinstance(value, int | float):
                acreage_value = float(value)
                return f"{acreage_value:.6f}".rstrip("0").rstrip(".")
            normalized = _strip_or_none(value)
            if normalized is None:
                return None
            match = re.search(r"-?\d+(?:\.\d+)?", normalized.replace(",", ""))
            if match is None:
                return None
            return match.group(0)
        acreage_value = float(value) / 43560
        return f"{acreage_value:.6f}".rstrip("0").rstrip(".")
    if transform == "acreage_text_or_square_feet":
        if value is None:
            return None
        if isinstance(value, int | float):
            acreage_value = float(value) / 43560
            return f"{acreage_value:.6f}".rstrip("0").rstrip(".")
        normalized = _strip_or_none(value)
        if normalized is None:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", normalized.replace(",", ""))
        if match is None:
            return None
        return match.group(0)
    if transform == "cad_mixed_area_to_acres":
        if value is None or str(value).strip() == "":
            return None
        normalized = _strip_or_none(value)
        if normalized is None:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", normalized.replace(",", ""))
        if match is None:
            return None
        numeric_value = float(match.group(0))
        if math.isnan(numeric_value) or math.isinf(numeric_value) or numeric_value < 0:
            return None

        source_name = (_strip_or_none(context.get("__source_name__")) or "").lower()
        source_identifier = re.sub(r"\D", "", match.group(0))
        parcel_identifiers = {
            re.sub(r"\D", "", identifier)
            for identifier in (
                _strip_or_none(context.get("HCAD_NUM")),
                _strip_or_none(context.get("LOWPARCELI")),
            )
            if identifier is not None
        }
        if source_identifier and source_identifier in parcel_identifiers:
            return None

        includes_acre_units = bool(re.search(r"\bac\b|acre", normalized, re.IGNORECASE))
        if "acreage" in source_name:
            if numeric_value > 100000:
                return None
            acreage_value = numeric_value
        elif includes_acre_units:
            acreage_value = numeric_value
        elif "statedarea" in source_name:
            acreage_value = numeric_value if numeric_value < 10000 else numeric_value / 43560
        else:
            acreage_value = numeric_value if numeric_value < 10000 else numeric_value / 43560

        if acreage_value > 1_000_000:
            return None
        return f"{acreage_value:.6f}".rstrip("0").rstrip(".")
    if transform == "extract_decimal":
        normalized = _strip_or_none(value)
        if normalized is None:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", normalized)
        if match is None:
            return None
        return match.group(0)
    if transform == "map_value":
        mapping = {
            str(key): mapped_value
            for key, mapped_value in (field_rule.options.get("map") or {}).items()
        }
        normalized = _strip_or_none(value)
        if normalized is None:
            return None
        if normalized in mapping:
            return mapping[normalized]
        if field_rule.options.get("preserve_unmapped", False):
            return normalized
        return field_rule.default
    if transform == "template":
        if field_rule.template is None:
            return value
        return field_rule.template.format_map(_SafeFormatDict(context))

    raise ConnectorConfigurationError(
        f"Unsupported transform `{field_rule.transform}` in connector field rule."
    )


def _extract_source_value(
    raw_record: dict[str, Any],
    source: str | list[str] | None,
) -> Any:
    if source is None:
        return None
    if isinstance(source, list):
        for candidate in source:
            value = _resolve_field(raw_record, candidate)
            if value is not None and str(value).strip() != "":
                return value
        return None
    return _resolve_field(raw_record, source)


def _resolve_field(raw_record: dict[str, Any], field_path: str) -> Any:
    current_value: Any = raw_record
    if isinstance(current_value, dict) and field_path in current_value:
        return current_value.get(field_path)
    for path_part in field_path.split("."):
        if isinstance(current_value, dict):
            current_value = current_value.get(path_part)
            continue
        if isinstance(current_value, list):
            if not path_part.isdigit():
                return None
            index = int(path_part)
            if index < 0 or index >= len(current_value):
                return None
            current_value = current_value[index]
            continue
        return None
    return current_value


def _normalize_rule_source(value: Any) -> str | list[str] | None:
    if value is None:
        return None
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    normalized = str(value).strip()
    return normalized or None


def _normalize_path(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    normalized = str(value).strip()
    return [part for part in normalized.split(".") if part]


def _normalize_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    normalized = str(value).strip()
    return [normalized] if normalized else []


def _normalize_optional_string(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _strip_or_none(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _coerce_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=UTC)
        return value.astimezone(UTC)
    if isinstance(value, (int, float)):
        timestamp_value = float(value)
        if timestamp_value > 10_000_000_000:
            timestamp_value /= 1000
        return datetime.fromtimestamp(timestamp_value, tz=UTC)
    try:
        normalized = str(value).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _default_load_strategy(source_id: str) -> str:
    normalized = source_id.strip().upper()
    if normalized == "PARCEL":
        return "parcel"
    if normalized == "ZONING":
        return "zoning"
    if normalized == "LISTING":
        return "market_listing"
    return "evidence"


def _slugify(value: str) -> str:
    return "-".join(
        chunk
        for chunk in "".join(
            character.lower() if character.isalnum() else " "
            for character in value
        ).split()
        if chunk
    )
